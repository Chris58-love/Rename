# ============================================================
# Universal Colab Rename App
# 单个 Colab 代码单元可直接运行的批量改名工具框架
#
# 本轮实现范围：
# - 工作区初始化与任务状态管理
# - 任意文件类型上传与保存
# - file_inventory 源文件索引
# - rename_state 在线编辑状态
# - ipywidgets 在线表格编辑主流程
# - 本地校验、执行改名、报告、ZIP 下载、规则字典自动改名模块入口占位
# ============================================================

import os
import re
import gc
import csv
import html as html_lib
import io
import json
import shutil
import time
import zipfile
from pathlib import Path

import pandas as pd
import ipywidgets as widgets
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from IPython.display import display, HTML, clear_output
from google.colab import files
from google.colab import output as colab_output


colab_output.enable_custom_widget_manager()


WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
}

MAX_STEM_LENGTH = 180

HISTORY_COLUMNS = [
    "时间戳", "任务批次ID", "原文件名", "原文件名主体", "原扩展名",
    "用户填写的新文件名主体", "清洗后的新文件名主体", "最终输出文件名",
    "是否保留原扩展名", "是否实际改名", "处理说明",
    "AI建议名", "AI建议原因", "AI置信度", "AI审核结果", "AI审核风险", "AI审核说明", "状态",
    "original_stem_norm", "final_stem_norm", "original_prefix_number",
    "final_prefix_number", "original_tokens", "final_tokens", "pattern_key",
]

DEFAULT_RULE_RENAME_MAP = {
    "13": "13_m_up",
    "14": "14_m_dw",
    "15": "15_r_up",
    "16": "16_r_dw",
    "17": "17_l_up",
    "18": "18_l_dw",
    "19": "19_next_area",
    "21": "21_two_min",
    "24": "24_brush_clean",
    "25": "25_bth_start",
    "26": "26_guide_l_u",
    "27": "27_guide_r_u",
    "28": "28_guide_l_d",
    "29": "29_guide_r_d",
    "30": "30_time_not_enough",
    "31": "31_half_achieve",
    "32": "32_30s_achieve",
    "33": "33_guide_m_u",
    "34": "34_guide_m_d",
    "78": "78_wake_up",
    "79": "79_encourage",
    "80": "80_slack_off",
}


def sanitize_filename_component(name: str) -> str:
    """清洗文件名主体，不处理扩展名。"""
    name = "" if name is None else str(name)
    name = re.sub(r"[\x00-\x1f\x7f]", "", name)
    name = re.sub(r'[\\/:\*\?"<>\|]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    name = name.strip(" .")

    if not name:
        name = "untitled"

    if name.upper() in WINDOWS_RESERVED_NAMES:
        name = f"{name}_"

    if len(name) > MAX_STEM_LENGTH:
        name = name[:MAX_STEM_LENGTH].rstrip(" .")

    if not name:
        name = "untitled"

    return name


def safe_remove(path: Path):
    try:
        if path.is_symlink() or path.is_file():
            path.unlink(missing_ok=True)
        elif path.is_dir():
            shutil.rmtree(path, ignore_errors=True)
    except Exception:
        pass


def split_filename(name: str):
    """保留扩展名原始大小写；无扩展名时 extension 为空。"""
    name = Path(str(name)).name
    if "." in name and not name.endswith(".") and not name.startswith("."):
        stem, extension = name.rsplit(".", 1)
        return stem, f".{extension}", True
    return name, "", False


def normalize_for_history(value: str) -> str:
    value = "" if value is None else str(value).lower().strip()
    value = re.sub(r"[\W_]+", " ", value, flags=re.UNICODE)
    return re.sub(r"\s+", " ", value).strip()


def extract_prefix_number(value: str) -> str:
    match = re.match(r"^\s*(\d+)", "" if value is None else str(value))
    return match.group(1) if match else ""


def tokenize_for_history(value: str) -> str:
    value = "" if value is None else str(value).lower()
    tokens = [token for token in re.split(r"[^0-9a-zA-Z\u4e00-\u9fff]+", value) if token]
    return " ".join(tokens)


def format_size(size_bytes: int) -> str:
    size_bytes = int(size_bytes or 0)
    units = ["B", "KB", "MB", "GB"]
    size = float(size_bytes)
    for unit in units:
        if size < 1024 or unit == units[-1]:
            return f"{size:.0f} {unit}" if unit == "B" else f"{size:.2f} {unit}"
        size /= 1024


def escape(value) -> str:
    return html_lib.escape("" if value is None else str(value))


def panel_html(title, body, tone="info"):
    colors = {
        "info": ("#eef4ff", "#1a73e8"),
        "success": ("#edf8f0", "#188038"),
        "warning": ("#fff7e8", "#b06000"),
        "error": ("#fdeeee", "#d93025"),
        "gray": ("#f7f8f9", "#5f6368"),
    }
    bg, fg = colors.get(tone, colors["info"])
    return f"""
    <div style="background:{bg};border-left:4px solid {fg};border-radius:8px;
                padding:12px 14px;margin:8px 0;color:#202124;line-height:1.7;">
      <div style="font-weight:700;color:{fg};margin-bottom:4px;">{escape(title)}</div>
      <div>{body}</div>
    </div>
    """


def dataframe_html(df: pd.DataFrame, title=None, max_rows=30):
    if df is None:
        df = pd.DataFrame()
    total = len(df)
    view_df = df.head(max_rows).copy() if max_rows and total > max_rows else df.copy()
    footer = ""
    if max_rows and total > max_rows:
        footer = f"<div style='color:#5f6368;font-size:12px;margin-top:8px;'>仅展示前 {max_rows} 行，完整状态已保留 {total} 行。</div>"
    title_html = f"<div style='font-weight:700;margin-bottom:8px;'>{escape(title)}</div>" if title else ""
    return f"""
    <div style="background:#fff;border:1px solid #e0e3e7;border-radius:8px;padding:12px;margin:8px 0;">
      {title_html}
      <style>
        .ucr-table table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
        .ucr-table th, .ucr-table td {{ border:1px solid #e8eaed; padding:7px 9px; text-align:left; vertical-align:top; }}
        .ucr-table th {{ background:#f8f9fa; font-weight:700; color:#202124; }}
        .ucr-table tr:nth-child(even) {{ background:#fcfcfd; }}
      </style>
      <div class="ucr-table">{view_df.to_html(index=False, escape=True, border=0)}</div>
      {footer}
    </div>
    """


def beautify_excel(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)
    wb.save(xlsx_path)


class DSV4FlashClient:
    def __init__(self, api_key, model="dsv4flash", base_url=None):
        self.api_key = api_key
        self.model = model or "dsv4flash"
        self.base_url = (base_url or "https://api.deepseek.com/v1").rstrip("/")

    def _chat_url(self):
        return f"{self.base_url}/chat/completions"

    def _extract_json_text(self, text):
        text = "" if text is None else str(text).strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
            text = re.sub(r"\s*```$", "", text).strip()
        first = text.find("{")
        last = text.rfind("}")
        if first >= 0 and last > first:
            text = text[first:last + 1]
        return text

    def chat_json(self, messages, temperature=0.2, timeout=60):
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "response_format": {"type": "json_object"},
        }
        response = requests.post(self._chat_url(), headers=headers, json=payload, timeout=timeout)
        if response.status_code >= 400 and "response_format" in response.text:
            payload.pop("response_format", None)
            response = requests.post(self._chat_url(), headers=headers, json=payload, timeout=timeout)
        if response.status_code >= 400:
            raise RuntimeError(f"AI 请求失败：HTTP {response.status_code}，{response.text[:500]}")
        data = response.json()
        content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        if not content:
            raise RuntimeError("模型返回空内容。")
        json_text = self._extract_json_text(content)
        try:
            return json.loads(json_text)
        except Exception as e:
            raise RuntimeError(f"JSON 解析失败：{e}；原始内容：{content[:500]}")

    def suggest_names(self, payload):
        system_prompt = (
            "你是文件批量改名助手。只根据文件名、用户规则和历史改名记录生成建议。"
            "不读取、不推测文件内容。suggested_stem 必须是不含扩展名的文件名主体。"
            "不得输出路径，不得输出非法文件名字符，不得修改 original_name。"
            "返回 items 数量应尽量与输入文件数量一致。无可靠建议时 suggested_stem 允许为空。"
            "必须返回严格 JSON，不得返回 Markdown，不得解释 JSON 之外的内容。"
        )
        user_prompt = (
            "请基于以下 payload 生成改名建议。返回格式必须为："
            '{"scene":"场景名称","items":[{"original_name":"...","suggested_stem":"...","reason":"...","confidence":"高/中/低"}]}'
            "\n\npayload:\n"
            + json.dumps(payload, ensure_ascii=False)
        )
        return self.chat_json([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ])

    def test_connection(self):
        return self.chat_json([
            {"role": "system", "content": "你只返回严格 JSON。"},
            {"role": "user", "content": '{"ping":"请返回 {\"ok\": true, \"message\": \"connected\"}"}'},
        ], timeout=30)


class UniversalColabRenameApp:
    def __init__(self):
        self.base_root = Path("/content/universal_colab_rename_tool")

        self.main_root = self.base_root / "main_task"
        self.main_upload_dir = self.main_root / "uploads"
        self.main_output_dir = self.main_root / "output"
        self.main_system_dir = self.main_root / "system"

        self.rule_root = self.base_root / "rule_task"
        self.rule_upload_dir = self.rule_root / "uploads"
        self.rule_output_dir = self.rule_root / "output"

        self.export_dir = self.base_root / "exports"
        self.history_dir = self.base_root / "history"
        self.history_xlsx_path = self.history_dir / "rename_history.xlsx"
        self.history_csv_path = self.history_dir / "rename_history.csv"

        self.file_inventory = []
        self.rename_state = {}
        self.table_widgets = {}
        self.app_stage = "init"
        self.is_dirty = False
        self.has_ai_suggestions = False
        self.has_ai_review = False
        self.has_history_suggestions = False
        self.has_template_imported = False
        self.upload_completed = False
        self.validation_passed = False
        self.latest_valid_plan_df = None
        self.latest_validation_result = None
        self.latest_download_path = None
        self.latest_download_label = ""
        self.latest_rule_download_path = None
        self.latest_rule_download_label = ""
        self.last_history_df = pd.DataFrame()
        self.last_suggestion_df = pd.DataFrame()
        self.ai_enabled = False
        self.ai_available = False
        self.ai_configured = False
        self.current_ai_mode = "local"
        self.has_uploaded_files = False
        self.latest_ai_suggestion_result = None
        self.latest_ai_review_result = None
        self.busy = False
        self._suspend_dirty = False

        self._build_ui()
        self._reset_workspace()
        self._reset_task_state(clear_outputs=True)
        self._set_status("等待上传文件", "info")
        self._set_progress(0, "等待上传文件", "info")
        self._refresh_sections()

    # --------------------------------------------------------
    # UI
    # --------------------------------------------------------
    def _build_ui(self):
        self.title_html = widgets.HTML(value="""
        <div style="padding:16px 18px;border-radius:10px;background:#1a73e8;color:#fff;
                    font-size:24px;font-weight:700;margin-bottom:10px;">
          Universal Colab Rename App
          <div style="font-size:13px;font-weight:400;margin-top:6px;opacity:.92;">
            通用 Colab 批量改名工具 · 在线表格主流程
          </div>
        </div>
        """)

        self.guide_html = widgets.HTML(value="""
        <div style="padding:14px 16px;border:1px solid #e5e7eb;border-radius:8px;
                    background:#fbfcff;line-height:1.8;color:#202124;margin-bottom:10px;">
          <b>使用说明</b><br>
          1. 点击“上传待改名文件”，可一次上传多个任意类型文件。<br>
          2. 上传后页面会生成源文件索引和在线可编辑表格。<br>
          3. 可逐行编辑，也可从表格批量粘贴新文件名或完整映射表。<br>
          4. 校验通过后执行批量改名，系统会复制文件到 output、生成报告并打包下载。<br>
          5. 当前不会解析文件内容，也不会直接修改上传的原文件。
        </div>
        """)

        self.status_html = widgets.HTML()
        self.progress_label = widgets.HTML()
        self.progress_bar = widgets.IntProgress(
            value=0, min=0, max=100, bar_style="info",
            layout=widgets.Layout(width="100%", height="22px"),
        )

        button_layout = widgets.Layout(width="210px", height="40px", margin="4px 6px 4px 0")
        wide_button_layout = widgets.Layout(width="240px", height="40px", margin="4px 6px 4px 0")

        self.btn_upload_files = widgets.Button(
            description="上传待改名文件", icon="upload", button_style="primary", layout=button_layout
        )
        self.btn_validate = widgets.Button(
            description="校验当前命名", icon="check", button_style="info", layout=button_layout
        )
        self.btn_execute = widgets.Button(
            description="执行批量改名", icon="play", button_style="success", layout=button_layout
        )
        self.btn_redownload_latest = widgets.Button(
            description="重新下载最近结果包", icon="download", button_style="success", layout=wide_button_layout
        )
        self.btn_use_original = widgets.Button(
            description="一键使用原文件名主体", icon="undo", button_style="warning", layout=wide_button_layout
        )
        self.btn_clear_selection = widgets.Button(
            description="清空选择", icon="square-o", button_style="", layout=button_layout
        )
        self.btn_reset = widgets.Button(
            description="结束并开启新任务", icon="refresh", button_style="danger", layout=wide_button_layout
        )
        self.btn_rule_placeholder = widgets.Button(
            description="规则字典自动改名模块", icon="book", button_style="", layout=wide_button_layout
        )
        self.btn_apply_paste_list = widgets.Button(
            description="应用粘贴的新文件名列表", icon="paste", button_style="primary",
            layout=widgets.Layout(width="230px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_apply_paste_mapping = widgets.Button(
            description="应用粘贴的完整映射表", icon="table", button_style="primary",
            layout=widgets.Layout(width="230px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_clear_paste = widgets.Button(
            description="清空粘贴内容", icon="trash", button_style="",
            layout=widgets.Layout(width="160px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_generate_history_suggestions = widgets.Button(
            description="根据历史记录生成推荐名", icon="magic", button_style="info",
            layout=widgets.Layout(width="240px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_apply_selected_history = widgets.Button(
            description="采用选中历史推荐", icon="check", button_style="warning",
            layout=widgets.Layout(width="200px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_apply_all_history = widgets.Button(
            description="采用全部历史推荐", icon="check-square", button_style="warning",
            layout=widgets.Layout(width="200px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_clear_history_suggestions = widgets.Button(
            description="清空历史推荐", icon="eraser", button_style="",
            layout=widgets.Layout(width="160px", height="38px", margin="4px 6px 4px 0"),
        )
        self.btn_refresh_history = widgets.Button(description="刷新历史记录", icon="refresh", layout=button_layout)
        self.btn_show_recent_history = widgets.Button(description="展示最近 30 条", icon="list", layout=button_layout)
        self.btn_show_all_history = widgets.Button(description="展示全部历史记录", icon="table", layout=wide_button_layout)
        self.btn_export_history = widgets.Button(description="导出历史记录", icon="download", button_style="success", layout=button_layout)
        self.btn_clear_history = widgets.Button(description="清空历史记录（不可恢复）", icon="trash", button_style="danger", layout=wide_button_layout)
        self.btn_generate_history_copy = widgets.Button(description="复制友好格式生成", icon="copy", layout=button_layout)
        self.btn_export_excel_template = widgets.Button(
            description="导出 Excel 模板", icon="file-excel-o", button_style="info", layout=button_layout
        )
        self.btn_import_excel_template = widgets.Button(
            description="上传 Excel 模板覆盖当前表格", icon="upload", button_style="warning", layout=wide_button_layout
        )
        self.btn_process_rule_files = widgets.Button(
            description="上传并按规则处理", icon="upload", button_style="success", layout=button_layout
        )
        self.btn_redownload_rule_latest = widgets.Button(
            description="重新下载规则模块结果包", icon="download", button_style="success", layout=wide_button_layout
        )
        self.btn_test_ai_connection = widgets.Button(
            description="测试连接", icon="plug", button_style="info", layout=button_layout
        )
        self.btn_generate_ai_suggestions = widgets.Button(
            description="生成 AI 建议名", icon="magic", button_style="primary", layout=button_layout
        )
        self.btn_apply_selected_ai = widgets.Button(
            description="采用选中 AI 建议", icon="check", button_style="warning", layout=button_layout
        )
        self.btn_apply_all_ai = widgets.Button(
            description="采用全部 AI 建议", icon="check-square", button_style="warning", layout=button_layout
        )
        self.btn_clear_ai_suggestions = widgets.Button(
            description="清空 AI 建议", icon="eraser", button_style="", layout=button_layout
        )
        self.btn_ai_review_names = widgets.Button(
            description="AI 审核当前命名", icon="search", button_style="info", layout=button_layout
        )
        self.btn_clear_ai_review = widgets.Button(
            description="清空 AI 审核", icon="eraser", button_style="", layout=button_layout
        )
        self.btn_ai_explain_issues = widgets.Button(
            description="AI 解释当前问题", icon="comments", button_style="info", layout=button_layout
        )
        self.btn_refresh_ai_guidance = widgets.Button(
            description="刷新 AI 指引", icon="refresh", button_style="info", layout=button_layout
        )
        self.btn_ai_qa_ask = widgets.Button(
            description="提问", icon="question", button_style="primary", layout=button_layout
        )
        self.btn_run_self_check = widgets.Button(
            description="运行基础自检", icon="check-circle", button_style="info", layout=button_layout
        )
        self.ai_mode_radio = widgets.RadioButtons(
            options=[
                ("本地模式，不启用 AI", "local"),
                ("DSV4Flash AI 辅助模式", "dsv4flash"),
            ],
            value="local",
            description="模式：",
            layout=widgets.Layout(width="360px"),
        )
        self.ai_mode_hint = widgets.HTML()

        self.btn_upload_files.on_click(self._on_upload_files_clicked)
        self.btn_validate.on_click(self._on_validate_clicked)
        self.btn_execute.on_click(self._on_execute_clicked)
        self.btn_redownload_latest.on_click(self._on_redownload_latest_clicked)
        self.btn_use_original.on_click(self._on_use_original_clicked)
        self.btn_clear_selection.on_click(self._on_clear_selection_clicked)
        self.btn_reset.on_click(self._on_reset_clicked)
        self.btn_rule_placeholder.on_click(self._on_rule_placeholder_clicked)
        self.btn_apply_paste_list.on_click(self._on_apply_paste_list_clicked)
        self.btn_apply_paste_mapping.on_click(self._on_apply_paste_mapping_clicked)
        self.btn_clear_paste.on_click(self._on_clear_paste_clicked)
        self.btn_generate_history_suggestions.on_click(self._on_generate_history_suggestions_clicked)
        self.btn_apply_selected_history.on_click(self._on_apply_selected_history_clicked)
        self.btn_apply_all_history.on_click(self._on_apply_all_history_clicked)
        self.btn_clear_history_suggestions.on_click(self._on_clear_history_suggestions_clicked)
        self.btn_refresh_history.on_click(self._on_refresh_history_clicked)
        self.btn_show_recent_history.on_click(self._on_show_recent_history_clicked)
        self.btn_show_all_history.on_click(self._on_show_all_history_clicked)
        self.btn_export_history.on_click(self._on_export_history_clicked)
        self.btn_clear_history.on_click(self._on_clear_history_clicked)
        self.btn_generate_history_copy.on_click(self._on_generate_history_copy_clicked)
        self.btn_export_excel_template.on_click(self._on_export_excel_template_clicked)
        self.btn_import_excel_template.on_click(self._on_import_excel_template_clicked)
        self.btn_process_rule_files.on_click(self._on_process_rule_files_clicked)
        self.btn_redownload_rule_latest.on_click(self._on_redownload_rule_latest_clicked)
        self.btn_test_ai_connection.on_click(self._on_test_ai_connection_clicked)
        self.btn_generate_ai_suggestions.on_click(self._on_generate_ai_suggestions_clicked)
        self.btn_apply_selected_ai.on_click(self._on_apply_selected_ai_clicked)
        self.btn_apply_all_ai.on_click(self._on_apply_all_ai_clicked)
        self.btn_clear_ai_suggestions.on_click(self._on_clear_ai_suggestions_clicked)
        self.btn_ai_review_names.on_click(self._on_ai_review_names_clicked)
        self.btn_clear_ai_review.on_click(self._on_clear_ai_review_clicked)
        self.btn_ai_explain_issues.on_click(self._on_ai_explain_issues_clicked)
        self.btn_refresh_ai_guidance.on_click(self._on_refresh_ai_guidance_clicked)
        self.btn_ai_qa_ask.on_click(self._on_ai_qa_ask_clicked)
        self.btn_run_self_check.on_click(self._on_run_self_check_clicked)
        self.ai_mode_radio.observe(self._on_ai_mode_changed, names="value")

        self.ai_mode_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:12px 0 6px;'>AI 辅助模式选择</h3>"),
            self.ai_mode_hint,
            self.ai_mode_radio,
        ])

        self.upload_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:12px 0 6px;'>文件上传区</h3>"),
            widgets.HBox([self.btn_upload_files, self.btn_reset]),
        ])

        self.summary_html = widgets.HTML()
        self.inventory_output = widgets.Output()

        self.table_section_title = widgets.HTML("<h3 style='margin:14px 0 6px;'>在线表格编辑区</h3>")
        self.table_display_limit = widgets.Dropdown(
            options=[("50", 50), ("100", 100), ("200", 200), ("全部", "all")],
            value=100,
            description="展示行数：",
            layout=widgets.Layout(width="220px"),
        )
        self.table_filter_text = widgets.Text(
            value="",
            description="过滤：",
            placeholder="按原文件名包含关键字过滤",
            layout=widgets.Layout(width="360px"),
        )
        self.table_display_limit.observe(lambda change: self._refresh_online_table_from_state(), names="value")
        self.table_filter_text.observe(lambda change: self._refresh_online_table_from_state(), names="value")
        self.table_help_html = widgets.HTML(value=panel_html(
            "在线表格",
            "上传文件后在这里编辑“选择 / 新文件名 / 是否保留原扩展名 / 采用AI建议”。AI 字段本轮仅保留占位，不影响手动编辑流程。",
            "gray",
        ))
        self.paste_list_text = widgets.Textarea(
            value="",
            placeholder="从 Excel/WPS/飞书表格复制一列新文件名后粘贴到这里，每行对应当前文件顺序的一行。",
            layout=widgets.Layout(width="100%", height="150px"),
        )
        self.paste_mapping_text = widgets.Textarea(
            value="",
            placeholder="从表格复制“原文件名 + 新文件名”两列或多列后粘贴到这里。支持 Tab 分隔、逗号分隔、首行表头。",
            layout=widgets.Layout(width="100%", height="150px"),
        )
        self.paste_result_output = widgets.Output()
        self.paste_section = widgets.VBox([
            widgets.HTML("<h4 style='margin:8px 0 6px;'>批量粘贴改名区</h4>"),
            widgets.HTML(panel_html(
                "批量粘贴",
                "左侧适合粘贴一列新文件名，按当前文件顺序覆盖；右侧适合粘贴完整映射表，按原文件名匹配更新。批量粘贴和逐行表格都以 rename_state 为唯一状态源。",
                "gray",
            )),
            widgets.HBox([
                widgets.VBox([
                    widgets.HTML("<b>粘贴新文件名列表</b>"),
                    self.paste_list_text,
                ], layout=widgets.Layout(width="50%", padding="0 8px 0 0")),
                widgets.VBox([
                    widgets.HTML("<b>粘贴完整映射表</b>"),
                    self.paste_mapping_text,
                ], layout=widgets.Layout(width="50%", padding="0 0 0 8px")),
            ]),
            widgets.HBox([self.btn_apply_paste_list, self.btn_apply_paste_mapping, self.btn_clear_paste]),
            self.paste_result_output,
        ])
        self.table_output = widgets.Output()

        self.template_result_output = widgets.Output()
        self.template_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>Excel 模板备用区域</h3>"),
            widgets.HTML(panel_html(
                "Excel 模板备用模式",
                "用于大批量离线编辑。模板导入只覆盖 rename_state，导入后仍需重新校验，不会直接执行改名。",
                "gray",
            )),
            widgets.HBox([self.btn_export_excel_template, self.btn_import_excel_template]),
            self.template_result_output,
        ])

        self.ai_api_key_text = widgets.Password(
            value="",
            description="API Key：",
            layout=widgets.Layout(width="520px"),
        )
        self.ai_base_url_text = widgets.Text(
            value="https://api.deepseek.com/v1",
            description="Base URL：",
            layout=widgets.Layout(width="520px"),
        )
        self.ai_model_text = widgets.Text(
            value="dsv4flash",
            description="模型名称：",
            layout=widgets.Layout(width="320px"),
        )
        self.ai_enabled_checkbox = widgets.Checkbox(value=False, description="启用 AI", indent=False)
        self.ai_enabled_checkbox.observe(self._on_ai_enabled_checkbox_changed, names="value")
        self.ai_status_output = widgets.Output()
        self.ai_settings_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>DSV4Flash 设置</h3>"),
            widgets.HBox([self.ai_enabled_checkbox, self.ai_model_text]),
            self.ai_api_key_text,
            self.ai_base_url_text,
            widgets.HBox([self.btn_test_ai_connection]),
            self.ai_status_output,
        ])
        self.ai_mode_section.children = tuple(list(self.ai_mode_section.children) + [self.ai_settings_section])

        self.ai_scene_dropdown = widgets.Dropdown(
            options=[
                "通用清洗型",
                "编号规则型",
                "电商数据文件型",
                "图片素材整理型",
                "音频/视频素材型",
                "文档归档型",
                "自定义规则型",
            ],
            value="通用清洗型",
            description="场景：",
            layout=widgets.Layout(width="320px"),
        )
        self.ai_scene_description = widgets.HTML()
        self.ai_user_instruction = widgets.Textarea(
            value="",
            description="补充要求：",
            placeholder="例如：保留中文核心含义，去掉临时、测试、copy 等词。",
            layout=widgets.Layout(width="100%", height="90px"),
        )
        self.ai_naming_format = widgets.Text(
            value="",
            description="命名格式：",
            placeholder="例如：平台_数据类型_对象_日期",
            layout=widgets.Layout(width="100%"),
        )
        self.ai_scene_params = widgets.Textarea(
            value="",
            description="场景参数：",
            layout=widgets.Layout(width="100%", height="150px"),
        )
        self.ai_result_output = widgets.Output()
        self.ai_review_instruction = widgets.Textarea(
            value="",
            description="审核要求：",
            placeholder="可选：例如重点检查月份格式、平台缩写、编号是否连续。",
            layout=widgets.Layout(width="100%", height="80px"),
        )
        self.ai_review_output = widgets.Output()
        self.ai_guidance_output = widgets.Output()
        self.ai_qa_question = widgets.Textarea(
            value="",
            description="问题：",
            placeholder="输入你对当前流程的疑问。",
            layout=widgets.Layout(width="100%", height="80px"),
        )
        self.ai_qa_output = widgets.Output()
        quick_questions = [
            "我下一步该做什么？",
            "为什么不能执行改名？",
            "如何处理重复文件名？",
            "AI 建议名怎么使用？",
            "是否可以不保留扩展名？",
            "Excel 模板模式和在线表格模式有什么区别？",
            "历史推荐和 AI 建议有什么区别？",
            "规则字典自动改名适合什么场景？",
            "为什么不能直接写扩展名？",
            "校验通过后还能修改吗？",
        ]
        self.ai_quick_question_buttons = []
        for question in quick_questions:
            btn = widgets.Button(description=question, layout=widgets.Layout(width="230px", height="34px", margin="3px"))
            btn.on_click(lambda _, q=question: self._on_quick_question_clicked(q))
            self.ai_quick_question_buttons.append(btn)
        self.ai_scene_dropdown.observe(self._on_ai_scene_changed, names="value")
        self.ai_assistant_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>DSV4Flash 场景化改名助手</h3>"),
            self.ai_scene_dropdown,
            self.ai_scene_description,
            self.ai_user_instruction,
            self.ai_naming_format,
            self.ai_scene_params,
            widgets.HBox([
                self.btn_generate_ai_suggestions,
                self.btn_apply_selected_ai,
                self.btn_apply_all_ai,
                self.btn_clear_ai_suggestions,
            ]),
            self.ai_result_output,
        ])
        self.ai_review_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>AI 命名审核区</h3>"),
            widgets.HTML(panel_html(
                "AI 审核",
                "审核命名风格、语义重复、编号异常、月份格式和误操作风险。审核只写入 AI审核字段，不修改新文件名，不替代本地硬校验。",
                "gray",
            )),
            self.ai_review_instruction,
            widgets.HBox([self.btn_ai_review_names, self.btn_clear_ai_review, self.btn_ai_explain_issues]),
            self.ai_review_output,
        ])
        self.ai_guidance_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>AI 当前指引</h3>"),
            widgets.HBox([self.btn_refresh_ai_guidance]),
            self.ai_guidance_output,
        ])
        self.ai_qa_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>新用户问答助手</h3>"),
            self.ai_qa_question,
            widgets.HBox([self.btn_ai_qa_ask]),
            widgets.HBox(self.ai_quick_question_buttons[:5]),
            widgets.HBox(self.ai_quick_question_buttons[5:]),
            self.ai_qa_output,
        ])
        self._on_ai_scene_changed(None)
        self.ai_accordion = widgets.Accordion(children=[widgets.VBox([
            self.ai_assistant_section,
            self.ai_review_section,
            self.ai_guidance_section,
            self.ai_qa_section,
        ])])
        self.ai_accordion.set_title(0, "DSV4Flash 设置与场景化改名助手")
        self.ai_accordion.selected_index = 0
        self.ai_api_key_text.observe(lambda change: self._refresh_ai_sections(), names="value")
        self.ai_base_url_text.observe(lambda change: self._refresh_ai_sections(), names="value")
        self.ai_model_text.observe(lambda change: self._refresh_ai_sections(), names="value")

        self.history_suggestion_output = widgets.Output()
        self.history_suggestion_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>历史推荐区</h3>"),
            widgets.HTML(panel_html(
                "历史推荐",
                "根据当前会话内保存的历史改名记录生成本地推荐。推荐只写入 AI建议名，不会自动覆盖新文件名。",
                "gray",
            )),
            widgets.HBox([
                self.btn_generate_history_suggestions,
                self.btn_apply_selected_history,
                self.btn_apply_all_history,
                self.btn_clear_history_suggestions,
            ]),
            self.history_suggestion_output,
        ])

        self.validation_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>校验结果区</h3>"),
            widgets.Output(),
        ])
        self.validation_output = self.validation_section.children[1]

        self.execute_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>执行与下载区</h3>"),
            widgets.HBox([
                self.btn_validate,
                self.btn_execute,
                self.btn_redownload_latest,
                self.btn_use_original,
                self.btn_clear_selection,
            ]),
            widgets.Output(),
        ])
        self.execute_output = self.execute_section.children[2]

        self.rule_zip_name_text = widgets.Text(
            value="",
            description="ZIP 名称：",
            placeholder="留空则自动生成 rule_renamed_files_时间戳.zip",
            layout=widgets.Layout(width="480px"),
        )
        self.rule_output = widgets.Output()
        self.rule_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>规则字典自动改名模块</h3>"),
            widgets.HTML(panel_html(
                "模块说明",
                "支持任意文件格式；只按文件名主体开头编号匹配规则。命中规则则替换为标准名称；未命中规则则保留原主体并清理默认后缀；默认保留原扩展名；若最终输出文件名冲突，系统会拦截。",
                "gray",
            )),
            widgets.HTML(dataframe_html(
                pd.DataFrame([{"编号": k, "标准名称": v} for k, v in DEFAULT_RULE_RENAME_MAP.items()]),
                title="当前规则字典预览",
                max_rows=80,
            )),
            widgets.HBox([self.rule_zip_name_text, self.btn_process_rule_files, self.btn_redownload_rule_latest]),
            self.rule_output,
        ])

        self.history_review_output = widgets.Output()
        self.history_copy_format = widgets.Dropdown(
            options=[
                ("两列表格文本", "table"),
                ("箭头格式", "arrow"),
                ("仅新旧主体", "stem"),
            ],
            value="table",
            description="格式：",
            layout=widgets.Layout(width="260px"),
        )
        self.history_copy_text = widgets.Textarea(
            value="",
            placeholder="点击“复制友好格式生成”后，这里会生成可复制文本。",
            layout=widgets.Layout(width="100%", height="180px"),
        )
        self.history_review_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>历史改名记录回顾</h3>"),
            widgets.HBox([
                self.btn_refresh_history,
                self.btn_show_recent_history,
                self.btn_show_all_history,
                self.btn_export_history,
                self.btn_clear_history,
            ]),
            self.history_review_output,
            widgets.HTML("<h3 style='margin:14px 0 6px;'>精简复制区</h3>"),
            widgets.HBox([self.history_copy_format, self.btn_generate_history_copy]),
            self.history_copy_text,
        ])

        self.log_output = widgets.Output(layout=widgets.Layout(max_height="260px", overflow_y="auto"))
        self.log_accordion = widgets.Accordion(children=[self.log_output])
        self.log_accordion.set_title(0, "日志区")
        self.log_accordion.selected_index = None
        self.self_check_output = widgets.Output()
        self.self_check_section = widgets.VBox([
            widgets.HTML("<h3 style='margin:14px 0 6px;'>全流程自检</h3>"),
            widgets.HTML(panel_html(
                "自检说明",
                "自检只检查目录、状态、控件挂载和必要方法，不执行真实改名、不触发下载、不读取文件内容。",
                "gray",
            )),
            widgets.HBox([self.btn_run_self_check]),
            self.self_check_output,
        ])
        self.self_check_accordion = widgets.Accordion(children=[self.self_check_section])
        self.self_check_accordion.set_title(0, "日志与自检")
        self.self_check_accordion.selected_index = None

        self.app_box = widgets.VBox([
            self.title_html,
            self.guide_html,
            widgets.HTML("<h3 style='margin:12px 0 6px;'>当前状态区</h3>"),
            self.status_html,
            widgets.HTML("<h3 style='margin:12px 0 6px;'>进度区</h3>"),
            self.progress_label,
            self.progress_bar,
            self.ai_mode_section,
            self.ai_accordion,
            self.upload_section,
            self.summary_html,
            self.inventory_output,
            self.table_section_title,
            widgets.HBox([self.table_display_limit, self.table_filter_text]),
            self.table_help_html,
            self.paste_section,
            self.table_output,
            self.template_section,
            self.history_suggestion_section,
            self.validation_section,
            self.execute_section,
            self.history_review_section,
            self.rule_section,
            widgets.HTML("<h3 style='margin:14px 0 6px;'>日志区</h3>"),
            self.log_accordion,
            self.self_check_accordion,
        ])

    def show(self):
        display(self.app_box)

    # --------------------------------------------------------
    # Workspace and state
    # --------------------------------------------------------
    def _reset_workspace(self):
        safe_remove(self.main_root)
        safe_remove(self.rule_root)
        safe_remove(self.export_dir)
        for directory in [
            self.main_upload_dir,
            self.main_output_dir,
            self.main_system_dir,
            self.rule_upload_dir,
            self.rule_output_dir,
            self.export_dir,
            self.history_dir,
        ]:
            directory.mkdir(parents=True, exist_ok=True)

    def _reset_task_state(self, clear_outputs=False):
        self.file_inventory = []
        self.rename_state = {}
        self.table_widgets = {}
        self.app_stage = "waiting_upload"
        self.is_dirty = False
        self.has_ai_suggestions = False
        self.has_ai_review = False
        self.has_history_suggestions = False
        self.has_template_imported = False
        self.upload_completed = False
        self.validation_passed = False
        self.latest_valid_plan_df = None
        self.latest_validation_result = None
        self.latest_download_path = None
        self.latest_download_label = ""
        self.latest_rule_download_path = None
        self.latest_rule_download_label = ""
        self.latest_ai_suggestion_result = None
        self.latest_ai_review_result = None
        self._suspend_dirty = False

        if clear_outputs:
            for output in [
                self.inventory_output,
                self.paste_result_output,
                self.table_output,
                self.template_result_output,
                self.history_suggestion_output,
                self.ai_status_output,
                self.ai_result_output,
                self.ai_review_output,
                self.ai_guidance_output,
                self.ai_qa_output,
                self.validation_output,
                self.execute_output,
                self.history_review_output,
                self.rule_output,
            ]:
                with output:
                    clear_output(wait=True)

        self.summary_html.value = ""
        self.paste_list_text.value = ""
        self.paste_mapping_text.value = ""
        self.history_copy_text.value = ""
        self.last_suggestion_df = pd.DataFrame()
        self._render_empty_table()
        self._refresh_sections()

    def _start_new_task(self):
        safe_remove(self.main_root)
        safe_remove(self.rule_root)
        safe_remove(self.export_dir)
        for directory in [
            self.main_upload_dir,
            self.main_output_dir,
            self.main_system_dir,
            self.rule_upload_dir,
            self.rule_output_dir,
            self.export_dir,
        ]:
            directory.mkdir(parents=True, exist_ok=True)
        self._reset_task_state(clear_outputs=True)

    # --------------------------------------------------------
    # Status, logs, buttons
    # --------------------------------------------------------
    def _set_status(self, message, level="info"):
        self.status_html.value = panel_html("当前状态", escape(message), level)

    def _set_progress(self, value, message, level="info"):
        style_map = {
            "info": "info",
            "success": "success",
            "warning": "warning",
            "error": "danger",
        }
        self.progress_bar.value = max(0, min(100, int(value)))
        self.progress_bar.bar_style = style_map.get(level, "info")
        self.progress_label.value = f"<div style='color:#5f6368;font-size:13px;margin-bottom:4px;'>{escape(message)}</div>"

    def _log(self, message):
        timestamp = time.strftime("%H:%M:%S")
        with self.log_output:
            print(f"[{timestamp}] {message}")

    def _set_stage(self, stage):
        self.app_stage = stage
        self.has_uploaded_files = bool(self.file_inventory)

    def _target_output(self, target):
        return {
            "main": self.validation_output,
            "validation": self.validation_output,
            "execute": self.execute_output,
            "template": self.template_result_output,
            "history": self.history_review_output,
            "history_suggestion": self.history_suggestion_output,
            "ai": self.ai_result_output,
            "ai_review": self.ai_review_output,
            "ai_guidance": self.ai_guidance_output,
            "ai_qa": self.ai_qa_output,
            "rule": self.rule_output,
            "self_check": self.self_check_output,
        }.get(target, self.validation_output)

    def _show_message(self, title, message, details=None, target="main", tone="info"):
        try:
            output = self._target_output(target)
            with output:
                clear_output(wait=True)
                display(HTML(panel_html(title, escape(message), tone)))
                if details is not None:
                    if isinstance(details, pd.DataFrame):
                        display(HTML(dataframe_html(details, title="明细", max_rows=100)))
                    elif isinstance(details, list):
                        display(HTML(dataframe_html(pd.DataFrame({"明细": details}), title="明细", max_rows=100)))
                    elif isinstance(details, dict):
                        display(HTML(dataframe_html(pd.DataFrame([details]), title="明细", max_rows=100)))
                    else:
                        display(HTML(panel_html("详细信息", escape(details), "gray")))
            self._log(f"{title}：{message}")
        except Exception as e:
            self._log(f"展示消息失败：{e}")

    def _show_error(self, title, message, details=None, target="main"):
        self._show_message(title, message, details=details, target=target, tone="error")

    def _show_warning(self, title, message, details=None, target="main"):
        self._show_message(title, message, details=details, target=target, tone="warning")

    def _show_success(self, title, message, details=None, target="main"):
        self._show_message(title, message, details=details, target=target, tone="success")

    def _refresh_buttons(self):
        uploaded = self.upload_completed
        self.has_uploaded_files = bool(self.file_inventory)
        self.ai_enabled = bool(getattr(self, "ai_enabled_checkbox", None) and self.ai_enabled_checkbox.value)
        self.current_ai_mode = "dsv4flash" if self.ai_enabled else "local"
        self.ai_configured = bool(getattr(self, "ai_api_key_text", None) and self.ai_api_key_text.value.strip())
        self.btn_upload_files.disabled = self.busy or uploaded
        self.btn_validate.disabled = self.busy or not uploaded
        self.btn_execute.disabled = self.busy or not uploaded or self.is_dirty or not self.validation_passed or self.latest_valid_plan_df is None
        self.btn_redownload_latest.disabled = self.busy or self.latest_download_path is None or not Path(self.latest_download_path).exists()
        self.btn_use_original.disabled = self.busy or not uploaded
        self.btn_clear_selection.disabled = self.busy or not uploaded
        self.btn_apply_paste_list.disabled = self.busy or not uploaded
        self.btn_apply_paste_mapping.disabled = self.busy or not uploaded
        self.btn_clear_paste.disabled = self.busy or not uploaded
        self.btn_export_excel_template.disabled = self.busy or not uploaded
        self.btn_import_excel_template.disabled = self.busy or not uploaded
        self.btn_generate_history_suggestions.disabled = self.busy or not uploaded
        self.btn_apply_selected_history.disabled = self.busy or not uploaded
        self.btn_apply_all_history.disabled = self.busy or not uploaded
        self.btn_clear_history_suggestions.disabled = self.busy or not uploaded
        self.btn_refresh_history.disabled = self.busy
        self.btn_show_recent_history.disabled = self.busy
        self.btn_show_all_history.disabled = self.busy
        self.btn_export_history.disabled = self.busy
        self.btn_clear_history.disabled = self.busy
        self.btn_generate_history_copy.disabled = self.busy
        self.btn_process_rule_files.disabled = self.busy
        self.btn_redownload_rule_latest.disabled = self.busy or self.latest_rule_download_path is None or not Path(self.latest_rule_download_path).exists()
        self.btn_test_ai_connection.disabled = self.busy or not self.ai_enabled
        ai_ready_for_files = uploaded and self.ai_enabled and self.ai_configured
        self.btn_generate_ai_suggestions.disabled = self.busy or not ai_ready_for_files
        self.btn_apply_selected_ai.disabled = self.busy or not uploaded or not self.ai_enabled
        self.btn_apply_all_ai.disabled = self.busy or not uploaded or not self.ai_enabled
        self.btn_clear_ai_suggestions.disabled = self.busy or not uploaded or not self.ai_enabled
        self.btn_ai_review_names.disabled = self.busy or not ai_ready_for_files
        self.btn_clear_ai_review.disabled = self.busy or not uploaded or not self.ai_enabled
        self.btn_ai_explain_issues.disabled = self.busy
        self.btn_refresh_ai_guidance.disabled = self.busy
        self.btn_ai_qa_ask.disabled = self.busy or not (self.ai_enabled and self.ai_configured)
        self.btn_run_self_check.disabled = self.busy
        self.btn_reset.disabled = self.busy
        self.btn_rule_placeholder.disabled = self.busy

    def _on_ai_mode_changed(self, change):
        self.current_ai_mode = change["new"]
        enabled = self.current_ai_mode == "dsv4flash"
        if self.ai_enabled_checkbox.value != enabled:
            self.ai_enabled_checkbox.value = enabled
        self._refresh_ai_sections()

    def _on_ai_enabled_checkbox_changed(self, change):
        enabled = bool(change["new"])
        self.ai_enabled = enabled
        mode = "dsv4flash" if enabled else "local"
        if self.ai_mode_radio.value != mode:
            self.ai_mode_radio.value = mode
        self._refresh_ai_sections()

    def _refresh_ai_sections(self):
        if not hasattr(self, "ai_enabled_checkbox"):
            return
        self.ai_enabled = bool(self.ai_enabled_checkbox.value)
        self.current_ai_mode = "dsv4flash" if self.ai_enabled else "local"
        self.has_uploaded_files = bool(self.file_inventory)
        self.ai_configured = bool(self.ai_api_key_text.value.strip())

        self.ai_mode_hint.value = panel_html(
            "AI 功能说明",
            "默认使用本地模式，不会触发任何网络请求。选择 DSV4Flash AI 辅助模式后，可填写 API Key 使用 AI 建议、AI 审核、动态指引和问答；AI 不会自动覆盖新文件名，也不会绕过本地校验。",
            "info" if self.ai_enabled else "gray",
        )

        self.ai_settings_section.layout.display = "" if self.ai_enabled else "none"
        self.ai_assistant_section.layout.display = ""
        self.ai_review_section.layout.display = ""
        self.ai_guidance_section.layout.display = ""
        self.ai_qa_section.layout.display = ""
        self.ai_accordion.selected_index = 0

        with self.ai_status_output:
            if not self.ai_enabled:
                clear_output(wait=True)
                display(HTML(panel_html("当前未启用 AI", "可在上方选择 DSV4Flash AI 辅助模式后填写 API Key。未启用 AI 时，本地功能、历史推荐、Excel 模板和规则字典模块均可正常使用。", "gray")))
            elif not self.ai_configured:
                clear_output(wait=True)
                display(HTML(panel_html("AI 已启用，等待 API Key", "请填写 API Key 后再测试连接或生成 AI 建议。API Key 仅保存在当前页面控件中，不写入文件、日志、报告或历史记录。", "warning")))

        if not self.has_uploaded_files:
            with self.ai_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("上传文件后可用", "上传文件后可使用 DSV4Flash 场景化改名建议。", "gray")))
            with self.ai_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("上传文件后可用", "上传文件后可使用 AI 命名审核。", "gray")))
        elif not self.ai_enabled:
            with self.ai_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("当前未启用 AI", "当前未启用 AI。可在上方启用 DSV4Flash AI 辅助后使用 AI 建议和 AI 审核。", "gray")))
            with self.ai_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("当前未启用 AI", "当前未启用 AI。可在上方启用 DSV4Flash AI 辅助后使用 AI 命名审核。", "gray")))
        elif not self.ai_configured:
            with self.ai_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("等待 API Key", "AI 已启用，但尚未填写 API Key。填写后即可生成 AI 建议名。", "warning")))
            with self.ai_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("等待 API Key", "AI 已启用，但尚未填写 API Key。填写后即可执行 AI 命名审核。", "warning")))

        guidance, steps = self._fixed_guidance(self._current_stage())
        with self.ai_guidance_output:
            clear_output(wait=True)
            note = "" if self.ai_enabled and self.ai_configured else "<br><span style='color:#5f6368;'>当前展示固定本地指引；填写 API Key 后可刷新动态 AI 指引。</span>"
            display(HTML(panel_html("当前步骤指引", escape(guidance) + note, "info")))
            if steps:
                display(HTML(dataframe_html(pd.DataFrame({"下一步": steps}), title="建议下一步", max_rows=20)))

        with self.ai_qa_output:
            clear_output(wait=True)
            if not self.ai_enabled:
                display(HTML(panel_html("新用户问答助手", "当前未启用 AI。快捷问题仍可使用本地兜底回答；启用 AI 并填写 API Key 后可获得动态回答。", "gray")))
            elif not self.ai_configured:
                display(HTML(panel_html("新用户问答助手", "请填写 API Key 后使用动态问答。", "warning")))
            else:
                display(HTML(panel_html("新用户问答助手", "可输入问题或点击快捷问题，AI 会基于当前流程状态回答。", "info")))

        self._refresh_buttons()

    def _show_main_error(self, title, message):
        self._set_status(message, "error")
        with self.validation_output:
            clear_output(wait=True)
            display(HTML(panel_html(title, escape(message), "error")))

    def _refresh_sections(self):
        upload_display = "" if self.upload_completed else "none"
        self.table_section_title.layout.display = ""
        self.paste_section.layout.display = upload_display
        self.template_section.layout.display = upload_display
        self.history_suggestion_section.layout.display = upload_display
        self.validation_section.layout.display = ""
        self.execute_section.layout.display = ""
        self.history_review_section.layout.display = ""
        self.rule_section.layout.display = ""
        self._refresh_ai_sections()
        self._refresh_buttons()

    def _run_basic_self_check(self):
        checks = []

        def add(name, ok, detail="", warning=False):
            checks.append({
                "检查项": name,
                "状态": "通过" if ok and not warning else ("警告" if warning else "失败"),
                "说明": detail,
                "建议修复方向": "" if ok else "检查控件初始化、目录创建或方法定义是否完整。",
            })

        dirs = [
            ("工作区目录", self.base_root),
            ("main_task/uploads", self.main_upload_dir),
            ("main_task/output", self.main_output_dir),
            ("exports", self.export_dir),
            ("history", self.history_dir),
        ]
        for name, path in dirs:
            if not path.exists():
                path.mkdir(parents=True, exist_ok=True)
            add(name, path.exists(), str(path))

        required_inventory_keys = {"original_name", "stem", "extension", "has_extension", "source_path", "size_bytes"}
        inv_ok = all(required_inventory_keys.issubset(set(item.keys())) for item in self.file_inventory) if self.file_inventory else True
        add("file_inventory 结构", inv_ok, f"当前记录数：{len(self.file_inventory)}")

        required_state_keys = {"selected", "new_stem", "keep_extension", "ai_suggested_stem", "ai_reason", "ai_confidence", "ai_review", "ai_review_risk_level", "ai_review_comment", "status"}
        state_ok = all(required_state_keys.issubset(set(state.keys())) for state in self.rename_state.values()) if self.rename_state else True
        add("rename_state 结构", state_ok, f"当前记录数：{len(self.rename_state)}")

        controls = [
            ("在线表格控件", "table_output"),
            ("批量粘贴控件", "paste_list_text"),
            ("校验按钮", "btn_validate"),
            ("执行按钮", "btn_execute"),
            ("Excel 模板按钮", "btn_export_excel_template"),
            ("历史推荐按钮", "btn_generate_history_suggestions"),
            ("AI 设置区", "ai_settings_section"),
            ("AI 建议区", "ai_assistant_section"),
            ("AI 审核区", "ai_review_section"),
            ("问答区", "ai_qa_section"),
            ("规则字典模块", "rule_section"),
        ]
        for label, attr in controls:
            add(label, hasattr(self, attr), attr)

        methods = [
            "_build_rename_plan",
            "_validate_current_state",
            "_execute_rename",
            "_build_excel_template",
            "_generate_history_based_suggestions",
            "_build_ai_payload",
            "_on_ai_review_names_clicked",
            "_build_rule_rename_plan",
            "_execute_rule_rename",
            "_refresh_ai_sections",
            "_refresh_buttons",
            "_run_basic_self_check",
        ]
        for method in methods:
            add(f"必要方法 {method}", callable(getattr(self, method, None)), method)

        add("最近结果包路径", self.latest_download_path is None or Path(self.latest_download_path).exists(), str(self.latest_download_path or "暂无"), warning=self.latest_download_path is not None and not Path(self.latest_download_path).exists())
        add("规则结果包路径", self.latest_rule_download_path is None or Path(self.latest_rule_download_path).exists(), str(self.latest_rule_download_path or "暂无"), warning=self.latest_rule_download_path is not None and not Path(self.latest_rule_download_path).exists())

        return pd.DataFrame(checks)

    def _on_run_self_check_clicked(self, _):
        try:
            result_df = self._run_basic_self_check()
            checklist = [
                "上传任意格式文件",
                "上传无扩展名文件",
                "上传中文文件名",
                "批量粘贴一列新文件名",
                "批量粘贴原文件名 + 新文件名映射",
                "导出 Excel 模板",
                "上传 Excel 模板覆盖当前表格",
                "生成历史推荐",
                "启用 AI 并生成 AI 建议",
                "AI 审核当前命名",
                "校验重复名",
                "执行批量改名",
                "下载 ZIP",
                "查看 rename_report.xlsx",
                "查看历史记录",
                "使用规则字典自动改名",
                "AI 失败时继续手动流程",
                "重置新任务后重新上传",
            ]
            with self.self_check_output:
                clear_output(wait=True)
                display(HTML(panel_html("基础自检完成", "自检不会执行真实改名、下载或读取文件内容。", "success")))
                display(HTML(dataframe_html(result_df, title="自检结果", max_rows=100)))
                display(HTML(dataframe_html(pd.DataFrame({"人工验收清单": checklist}), title="人工验收清单", max_rows=30)))
        except Exception as e:
            self._show_error("自检失败", str(e), target="self_check")

    # --------------------------------------------------------
    # Upload and inventory
    # --------------------------------------------------------
    def _on_upload_files_clicked(self, _):
        if self.busy:
            return

        self.busy = True
        self._refresh_buttons()

        try:
            self._start_new_task()
            self._set_status("正在上传文件", "info")
            self._set_progress(10, "等待选择上传文件", "info")
            self._log("等待用户选择待改名文件。")

            previous_cwd = os.getcwd()
            os.chdir("/content")
            uploaded = files.upload()
            os.chdir(previous_cwd)

            if not uploaded:
                self._set_status("未上传文件，请重新点击上传待改名文件。", "warning")
                self._set_progress(0, "等待上传文件", "warning")
                self._log("本次没有上传任何文件。")
                return

            self._save_uploaded_files(uploaded)
            del uploaded
            gc.collect()

            self._build_file_inventory()
            self._build_initial_rename_state()
            self.upload_completed = True
            self.validation_passed = False
            self.is_dirty = True
            self._set_stage("uploaded")

            self._render_upload_summary()
            self._render_file_inventory_preview()
            self._render_online_table()

            self._set_status("在线表格已生成", "success")
            self._set_progress(100, "文件上传完成，在线表格已生成", "success")
            self._refresh_sections()
            self._log(f"文件上传完成，共 {len(self.file_inventory)} 个文件。")

        except Exception as e:
            self._show_main_error("上传失败", str(e))
            self._set_progress(0, "上传失败", "error")
            self._log(f"上传失败：{e}")

        finally:
            self.busy = False
            self._refresh_sections()
            try:
                os.chdir("/content")
            except Exception:
                pass

    def _save_uploaded_files(self, uploaded):
        for original_name, content in uploaded.items():
            safe_name = Path(original_name).name
            if not safe_name:
                safe_name = "untitled"

            target_path = self.main_upload_dir / safe_name
            if target_path.exists():
                safe_remove(target_path)

            with open(target_path, "wb") as f:
                f.write(content)

            temp_colab_path = Path("/content") / safe_name
            if temp_colab_path.exists() and temp_colab_path != target_path:
                safe_remove(temp_colab_path)

    def _build_file_inventory(self):
        inventory = []
        for file_path in sorted(self.main_upload_dir.iterdir(), key=lambda p: p.name.lower()):
            if not file_path.is_file():
                continue
            stem, extension, has_extension = split_filename(file_path.name)
            inventory.append({
                "original_name": file_path.name,
                "stem": stem,
                "extension": extension,
                "has_extension": has_extension,
                "source_path": str(file_path),
                "size_bytes": file_path.stat().st_size,
            })
        self.file_inventory = inventory

    def _build_initial_rename_state(self):
        state = {}
        for item in self.file_inventory:
            state[item["original_name"]] = {
                "selected": False,
                "new_stem": sanitize_filename_component(item["stem"]),
                "keep_extension": True,
                "ai_suggested_stem": "",
                "ai_reason": "",
                "ai_confidence": "",
                "use_ai_suggestion": False,
                "ai_review": "",
                "ai_review_risk_level": "",
                "ai_review_comment": "",
                "remark": "",
                "status": "待编辑",
            }
        self.rename_state = state

    def _render_upload_summary(self):
        total_size = sum(item["size_bytes"] for item in self.file_inventory)
        with_ext = sum(1 for item in self.file_inventory if item["has_extension"])
        without_ext = len(self.file_inventory) - with_ext
        self.summary_html.value = f"""
        <div style="display:flex;gap:10px;flex-wrap:wrap;margin:10px 0;">
          <div style="border:1px solid #e0e3e7;border-radius:8px;padding:12px;min-width:150px;background:#fff;">
            <div style="color:#5f6368;font-size:12px;">已上传文件数</div>
            <div style="font-size:22px;font-weight:700;color:#1a73e8;">{len(self.file_inventory)}</div>
          </div>
          <div style="border:1px solid #e0e3e7;border-radius:8px;padding:12px;min-width:150px;background:#fff;">
            <div style="color:#5f6368;font-size:12px;">总大小</div>
            <div style="font-size:22px;font-weight:700;color:#188038;">{format_size(total_size)}</div>
          </div>
          <div style="border:1px solid #e0e3e7;border-radius:8px;padding:12px;min-width:150px;background:#fff;">
            <div style="color:#5f6368;font-size:12px;">有扩展名 / 无扩展名</div>
            <div style="font-size:22px;font-weight:700;color:#b06000;">{with_ext} / {without_ext}</div>
          </div>
        </div>
        """

    def _render_file_inventory_preview(self):
        rows = []
        for item in self.file_inventory:
            rows.append({
                "原文件名": item["original_name"],
                "文件名主体": item["stem"],
                "原扩展名": item["extension"],
                "是否有扩展名": "是" if item["has_extension"] else "否",
                "大小": format_size(item["size_bytes"]),
                "保存路径": item["source_path"],
            })
        df = pd.DataFrame(rows)
        with self.inventory_output:
            clear_output(wait=True)
            display(HTML(dataframe_html(df, title="file_inventory 源文件索引预览", max_rows=30)))

    # --------------------------------------------------------
    # Online table
    # --------------------------------------------------------
    def _render_empty_table(self):
        with self.table_output:
            clear_output(wait=True)
            display(HTML(panel_html("在线表格编辑区", "上传文件后会在这里生成可编辑表格。", "gray")))

    def _render_online_table(self):
        self.table_widgets = {}
        self._suspend_dirty = True

        header_style = "font-weight:700;color:#202124;background:#f8f9fa;padding:8px;border-bottom:1px solid #e8eaed;"
        cell_style = "padding:6px;border-bottom:1px solid #eef0f2;"

        header = widgets.GridBox(
            children=[
                widgets.HTML(f"<div style='{header_style}'>选择</div>"),
                widgets.HTML(f"<div style='{header_style}'>原文件名</div>"),
                widgets.HTML(f"<div style='{header_style}'>原扩展名</div>"),
                widgets.HTML(f"<div style='{header_style}'>新文件名（不含扩展名）</div>"),
                widgets.HTML(f"<div style='{header_style}'>是否保留原扩展名</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI建议名</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI建议原因</div>"),
                widgets.HTML(f"<div style='{header_style}'>采用AI建议</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI置信度</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI审核结果</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI审核风险</div>"),
                widgets.HTML(f"<div style='{header_style}'>AI审核说明</div>"),
                widgets.HTML(f"<div style='{header_style}'>状态</div>"),
            ],
            layout=widgets.Layout(
                grid_template_columns="70px minmax(180px,1.2fr) 100px minmax(220px,1.4fr) 150px 150px 180px 120px 100px 150px 110px 220px 110px",
                width="100%",
            ),
        )

        filter_text = self.table_filter_text.value.strip().lower() if hasattr(self, "table_filter_text") else ""
        display_items = [
            item for item in self.file_inventory
            if not filter_text or filter_text in item["original_name"].lower()
        ]
        total_items = len(display_items)
        limit_value = self.table_display_limit.value if hasattr(self, "table_display_limit") else 100
        if limit_value != "all":
            display_items = display_items[:int(limit_value)]

        rows = []
        for item in display_items:
            original_name = item["original_name"]
            state = self.rename_state[original_name]

            selected = widgets.Checkbox(value=bool(state["selected"]), indent=False, layout=widgets.Layout(width="60px"))
            original = widgets.HTML(f"<div style='{cell_style}'>{escape(original_name)}</div>")
            extension = widgets.HTML(f"<div style='{cell_style}'>{escape(item['extension'])}</div>")
            new_stem = widgets.Text(value=state["new_stem"], layout=widgets.Layout(width="100%"))
            keep_extension = widgets.Checkbox(value=bool(state["keep_extension"]), indent=False, layout=widgets.Layout(width="80px"))
            ai_name = widgets.HTML(f"<div style='{cell_style}'>{escape(state['ai_suggested_stem'])}</div>")
            ai_reason = widgets.HTML(f"<div style='{cell_style}'>{escape(state['ai_reason'])}</div>")
            use_ai = widgets.Checkbox(value=bool(state.get("use_ai_suggestion", False)), indent=False, layout=widgets.Layout(width="80px"))
            ai_confidence = widgets.HTML(f"<div style='{cell_style}'>{escape(state.get('ai_confidence', ''))}</div>")
            ai_review = widgets.HTML(f"<div style='{cell_style}'>{escape(state['ai_review'])}</div>")
            ai_review_risk = widgets.HTML(f"<div style='{cell_style}'>{escape(state.get('ai_review_risk_level', ''))}</div>")
            ai_review_comment = widgets.HTML(f"<div style='{cell_style}'>{escape(state.get('ai_review_comment', ''))}</div>")
            status = widgets.HTML(f"<div style='{cell_style}'>{escape(state['status'])}</div>")

            for control in [selected, new_stem, keep_extension, use_ai]:
                control.observe(self._on_table_control_changed, names="value")

            self.table_widgets[original_name] = {
                "selected": selected,
                "new_stem": new_stem,
                "keep_extension": keep_extension,
                "use_ai_suggestion": use_ai,
            }

            rows.append(widgets.GridBox(
                children=[
                    widgets.Box([selected], layout=widgets.Layout(padding="6px", border_bottom="1px solid #eef0f2")),
                    original,
                    extension,
                    widgets.Box([new_stem], layout=widgets.Layout(padding="6px", border_bottom="1px solid #eef0f2")),
                    widgets.Box([keep_extension], layout=widgets.Layout(padding="6px", border_bottom="1px solid #eef0f2")),
                    ai_name,
                    ai_reason,
                    widgets.Box([use_ai], layout=widgets.Layout(padding="6px", border_bottom="1px solid #eef0f2")),
                    ai_confidence,
                    ai_review,
                    ai_review_risk,
                    ai_review_comment,
                    status,
                ],
                layout=widgets.Layout(
                    grid_template_columns="70px minmax(180px,1.2fr) 100px minmax(220px,1.4fr) 150px 150px 180px 120px 100px 150px 110px 220px 110px",
                    width="100%",
                ),
            ))

        notice = widgets.HTML(panel_html(
            "展示提示",
            f"当前仅展示 <b>{len(display_items)}</b> 行，过滤后共有 <b>{total_items}</b> 行，实际会处理全部 <b>{len(self.file_inventory)}</b> 个文件。搜索和展示行数只影响页面渲染，不影响批量粘贴、模板导入、校验和执行。",
            "info" if len(display_items) < len(self.file_inventory) else "gray",
        ))

        table = widgets.VBox(
            [notice, header] + rows,
            layout=widgets.Layout(
                border="1px solid #e0e3e7",
                border_radius="8px",
                overflow_x="auto",
                background="#fff",
            ),
        )

        with self.table_output:
            clear_output(wait=True)
            display(table)

        self._suspend_dirty = False
        self._log("在线表格已渲染。")

    def _read_online_table_to_state(self):
        for original_name, controls in self.table_widgets.items():
            if original_name not in self.rename_state:
                continue
            self.rename_state[original_name]["selected"] = bool(controls["selected"].value)
            self.rename_state[original_name]["new_stem"] = str(controls["new_stem"].value)
            self.rename_state[original_name]["keep_extension"] = bool(controls["keep_extension"].value)
            self.rename_state[original_name]["use_ai_suggestion"] = bool(controls["use_ai_suggestion"].value)

    def _refresh_online_table_from_state(self):
        if not self.file_inventory:
            self._render_empty_table()
            return
        self._render_online_table()

    def _on_table_control_changed(self, _):
        if self._suspend_dirty:
            return
        self._read_online_table_to_state()
        self._mark_dirty()

    def _mark_dirty(self):
        self.is_dirty = True
        if self.app_stage not in {"pasted", "template_imported", "history_suggested", "ai_suggested", "ai_reviewed"}:
            self._set_stage("editing")
        self.validation_passed = False
        self.latest_valid_plan_df = None
        self.latest_validation_result = None
        self.btn_execute.disabled = True
        self._set_status("表格已修改，等待重新校验", "warning")
        self._set_progress(70, "表格已修改，等待重新校验", "warning")

    # --------------------------------------------------------
    # Batch paste
    # --------------------------------------------------------
    def _strip_same_original_extension(self, pasted_name: str, item: dict) -> str:
        value = "" if pasted_name is None else str(pasted_name).strip()
        extension = item.get("extension", "")
        if extension and value.lower().endswith(extension.lower()):
            value = value[:-len(extension)].rstrip(" .")
        return value

    def _show_paste_report(self, title, stats, unmatched=None, errors=None, tone="info"):
        unmatched = unmatched or []
        errors = errors or []
        details = [
            f"成功更新数量：<b>{stats.get('updated', 0)}</b>",
            f"跳过空行数量：<b>{stats.get('skipped_empty', 0)}</b>",
            f"多余行数量：<b>{stats.get('extra_rows', 0)}</b>",
            f"未匹配原文件名数量：<b>{stats.get('unmatched', 0)}</b>",
        ]

        detail_html = "<br>".join(details)
        if unmatched:
            rows = "".join(
                f"<tr><td>{escape(row.get('row_no', ''))}</td><td>{escape(row.get('original_name', ''))}</td><td>{escape(row.get('new_name', ''))}</td></tr>"
                for row in unmatched[:50]
            )
            detail_html += f"""
            <div style="margin-top:10px;font-weight:700;">未匹配明细</div>
            <table style="border-collapse:collapse;width:100%;font-size:12px;margin-top:4px;">
              <tr><th style="border:1px solid #ddd;padding:4px;">行号</th><th style="border:1px solid #ddd;padding:4px;">原文件名</th><th style="border:1px solid #ddd;padding:4px;">新文件名</th></tr>
              {rows}
            </table>
            """
            if len(unmatched) > 50:
                detail_html += f"<div style='color:#5f6368;font-size:12px;margin-top:4px;'>仅展示前 50 条，完整未匹配数量为 {len(unmatched)}。</div>"

        if errors:
            error_items = "".join(f"<li>{escape(err)}</li>" for err in errors[:50])
            detail_html += f"<div style='margin-top:10px;font-weight:700;'>错误明细</div><ul>{error_items}</ul>"
            if len(errors) > 50:
                detail_html += f"<div style='color:#5f6368;font-size:12px;'>仅展示前 50 条错误。</div>"

        with self.paste_result_output:
            clear_output(wait=True)
            display(HTML(panel_html(title, detail_html, tone)))

    def _parse_mapping_rows(self, raw_text: str):
        lines = [line for line in str(raw_text or "").splitlines() if line.strip()]
        if not lines:
            return [], []

        rows = []
        errors = []
        for index, line in enumerate(lines, start=1):
            try:
                if "\t" in line:
                    cols = [col.strip() for col in line.split("\t")]
                else:
                    cols = next(csv.reader(io.StringIO(line)))
                    cols = [col.strip() for col in cols]
                rows.append({"row_no": index, "cols": cols})
            except Exception as e:
                errors.append(f"第 {index} 行解析失败：{e}")
        return rows, errors

    def _detect_mapping_columns(self, rows):
        if not rows:
            return 0, 1, False

        header = [str(col).strip() for col in rows[0]["cols"]]
        normalized = [col.replace("(", "（").replace(")", "）").replace(" ", "") for col in header]

        original_candidates = ["原文件名", "original_name", "originalname"]
        new_candidates = ["新文件名（不含扩展名）", "新文件名", "new_stem", "newstem", "new_name", "newname"]

        original_col = None
        new_col = None
        for idx, name in enumerate(normalized):
            lower_name = name.lower()
            if original_col is None and (name in original_candidates or lower_name in original_candidates):
                original_col = idx
            if new_col is None and (name in new_candidates or lower_name in new_candidates):
                new_col = idx

        if original_col is not None and new_col is not None:
            return original_col, new_col, True

        return 0, 1, False

    def _mapping_has_enough_columns(self, rows, original_col, new_col, has_header):
        data_rows = rows[1:] if has_header else rows
        return any(len(row["cols"]) > max(original_col, new_col) for row in data_rows)

    def _on_apply_paste_list_clicked(self, _):
        if not self.upload_completed:
            return

        self._read_online_table_to_state()
        raw_text = self.paste_list_text.value
        lines = str(raw_text or "").splitlines()
        stats = {"updated": 0, "skipped_empty": 0, "extra_rows": 0, "unmatched": 0}
        errors = []

        if not lines:
            self._show_paste_report("粘贴新文件名列表", stats, errors=["粘贴内容为空。"], tone="warning")
            return

        for index, line in enumerate(lines):
            if index >= len(self.file_inventory):
                if str(line).strip():
                    stats["extra_rows"] += 1
                else:
                    stats["skipped_empty"] += 1
                continue

            item = self.file_inventory[index]
            value = self._strip_same_original_extension(line, item)
            if not value:
                stats["skipped_empty"] += 1
                continue

            self.rename_state[item["original_name"]]["new_stem"] = value
            stats["updated"] += 1

        self._refresh_online_table_from_state()
        self._set_stage("pasted")
        self._mark_dirty()
        tone = "success" if stats["updated"] else "warning"
        self._show_paste_report("已应用粘贴的新文件名列表", stats, errors=errors, tone=tone)
        self._log(f"已应用一列新文件名粘贴：更新 {stats['updated']} 条。")

    def _on_apply_paste_mapping_clicked(self, _):
        if not self.upload_completed:
            return

        self._read_online_table_to_state()
        rows, errors = self._parse_mapping_rows(self.paste_mapping_text.value)
        stats = {"updated": 0, "skipped_empty": 0, "extra_rows": 0, "unmatched": 0}
        unmatched = []

        if not rows:
            self._show_paste_report("粘贴完整映射表", stats, errors=errors + ["粘贴内容为空。"], tone="warning")
            return

        original_col, new_col, has_header = self._detect_mapping_columns(rows)
        data_rows = rows[1:] if has_header else rows
        inventory_by_name = {item["original_name"]: item for item in self.file_inventory}

        if not self._mapping_has_enough_columns(rows, original_col, new_col, has_header):
            errors.append("未找到可用的两列数据。若只粘贴一列，请使用左侧“粘贴新文件名列表”。")

        for row in data_rows:
            cols = row["cols"]
            if len(cols) <= max(original_col, new_col):
                errors.append(f"第 {row['row_no']} 行列数不足，无法读取原文件名和新文件名。")
                continue

            original_name = cols[original_col].strip()
            new_name = cols[new_col].strip()
            if not original_name and not new_name:
                stats["skipped_empty"] += 1
                continue
            if not new_name:
                stats["skipped_empty"] += 1
                continue

            item = inventory_by_name.get(original_name)
            if not item:
                stats["unmatched"] += 1
                unmatched.append({
                    "row_no": row["row_no"],
                    "original_name": original_name,
                    "new_name": new_name,
                })
                continue

            new_stem = self._strip_same_original_extension(new_name, item)
            if not new_stem:
                stats["skipped_empty"] += 1
                continue

            self.rename_state[original_name]["new_stem"] = new_stem
            stats["updated"] += 1

        self._refresh_online_table_from_state()
        self._set_stage("pasted")
        self._mark_dirty()
        tone = "success" if stats["updated"] and not unmatched and not errors else "warning"
        self._show_paste_report("已应用粘贴的完整映射表", stats, unmatched=unmatched, errors=errors, tone=tone)
        self._log(f"已应用完整映射表粘贴：更新 {stats['updated']} 条，未匹配 {stats['unmatched']} 条。")

    def _on_clear_paste_clicked(self, _):
        self.paste_list_text.value = ""
        self.paste_mapping_text.value = ""
        with self.paste_result_output:
            clear_output(wait=True)
        self._log("已清空批量粘贴内容。")

    # --------------------------------------------------------
    # Excel template fallback mode
    # --------------------------------------------------------
    def _parse_bool_value(self, value, default=None):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return default
        text = str(value).strip()
        if text == "":
            return default
        lower = text.lower()
        if lower in {"true", "1", "yes", "y", "是", "对", "勾选"}:
            return True
        if lower in {"false", "0", "no", "n", "否", "不", "未勾选"}:
            return False
        return default

    def _template_rows_df(self):
        rows = []
        for item in self.file_inventory:
            state = self.rename_state.get(item["original_name"], {})
            rows.append({
                "选择": bool(state.get("selected", False)),
                "原文件名": item["original_name"],
                "原扩展名": item["extension"],
                "新文件名（不含扩展名）": state.get("new_stem", ""),
                "是否保留原扩展名": bool(state.get("keep_extension", True)),
                "AI建议名": state.get("ai_suggested_stem", ""),
                "AI建议原因": state.get("ai_reason", ""),
                "AI置信度": state.get("ai_confidence", ""),
                "采用AI建议": bool(state.get("use_ai_suggestion", False)),
                "AI审核结果": state.get("ai_review", ""),
                "AI审核风险": state.get("ai_review_risk_level", ""),
                "AI审核说明": state.get("ai_review_comment", ""),
                "状态": state.get("status", "待编辑"),
                "备注": state.get("remark", ""),
            })
        return pd.DataFrame(rows)

    def _build_excel_template(self):
        self._read_online_table_to_state()
        self.main_system_dir.mkdir(parents=True, exist_ok=True)
        template_path = self.main_system_dir / "rename_template.xlsx"
        template_df = self._template_rows_df()
        instructions_df = pd.DataFrame({
            "项目": [
                "原文件名",
                "新文件名（不含扩展名）",
                "是否保留原扩展名",
                "选择 / 采用AI建议",
                "导入规则",
                "执行规则",
            ],
            "说明": [
                "用于回传匹配当前任务文件，请不要修改。",
                "主要编辑列。可留空，留空时导入不会覆盖当前在线表格中的 new_stem。",
                "支持 TRUE/FALSE、true/false、是/否、1/0、Y/N；留空时保留当前值。",
                "支持 TRUE/FALSE、是/否、1/0、Y/N。",
                "模板可以只保留部分文件行；缺失文件会保持当前在线状态不变。",
                "模板导入后不会直接执行改名，仍需点击“校验当前命名”。",
            ],
        })
        with pd.ExcelWriter(template_path, engine="openpyxl") as writer:
            template_df.to_excel(writer, index=False, sheet_name="rename_template")
            instructions_df.to_excel(writer, index=False, sheet_name="填写说明")
        beautify_excel(template_path)
        return template_path

    def _normalize_template_columns(self, df):
        alias_map = {
            "原文件名": ["原文件名", "original_name"],
            "新文件名": ["新文件名（不含扩展名）", "新文件名", "new_stem"],
            "是否保留原扩展名": ["是否保留原扩展名", "保留扩展名", "keep_extension"],
            "选择": ["选择", "selected"],
            "采用AI建议": ["采用AI建议", "use_ai_suggestion"],
            "AI建议名": ["AI建议名", "ai_suggested_stem"],
            "AI建议原因": ["AI建议原因", "ai_reason"],
            "AI置信度": ["AI置信度", "ai_confidence"],
            "AI审核结果": ["AI审核结果", "ai_review"],
            "AI审核风险": ["AI审核风险", "ai_review_risk_level"],
            "AI审核说明": ["AI审核说明", "ai_review_comment"],
            "状态": ["状态", "status"],
            "备注": ["备注", "remark"],
        }
        normalized = {}
        cols = {str(col).strip(): col for col in df.columns}
        lower_cols = {str(col).strip().lower(): col for col in df.columns}
        for canonical, aliases in alias_map.items():
            for alias in aliases:
                if alias in cols:
                    normalized[canonical] = cols[alias]
                    break
                if alias.lower() in lower_cols:
                    normalized[canonical] = lower_cols[alias.lower()]
                    break
        return normalized

    def _import_template_df_to_state(self, df):
        if df.empty:
            return None, {"ok": False, "errors": ["模板为空。"]}
        cols = self._normalize_template_columns(df)
        errors = []
        if "原文件名" not in cols:
            errors.append("缺少原文件名列。")
        if "新文件名" not in cols:
            errors.append("缺少新文件名列。")
        if errors:
            return None, {"ok": False, "errors": errors}

        original_col = cols["原文件名"]
        duplicate_names = df[original_col].dropna().astype(str).str.strip()
        duplicate_names = duplicate_names[duplicate_names != ""]
        dup = duplicate_names[duplicate_names.duplicated(keep=False)]
        if not dup.empty:
            dup_values = sorted(set(dup.tolist()))
            return None, {"ok": False, "errors": [f"模板内原文件名重复：{', '.join(dup_values[:20])}"]}

        inventory_names = {item["original_name"] for item in self.file_inventory}
        new_state = {key: dict(value) for key, value in self.rename_state.items()}
        unmatched = []
        updated = 0
        skipped_empty = 0
        parsed_rows = 0

        for row_index, row in df.iterrows():
            original_name = "" if pd.isna(row.get(original_col)) else str(row.get(original_col)).strip()
            if not original_name:
                skipped_empty += 1
                continue
            parsed_rows += 1
            if original_name not in inventory_names:
                unmatched.append({"行号": row_index + 2, "原文件名": original_name, "说明": "当前任务不存在该文件"})
                continue
            state = new_state.get(original_name, {})

            new_name_col = cols.get("新文件名")
            raw_new_name = row.get(new_name_col) if new_name_col else ""
            if raw_new_name is None or (isinstance(raw_new_name, float) and pd.isna(raw_new_name)) or str(raw_new_name).strip() == "":
                skipped_empty += 1
            else:
                state["new_stem"] = str(raw_new_name).strip()
                updated += 1

            for canonical, state_key in [
                ("选择", "selected"),
                ("是否保留原扩展名", "keep_extension"),
                ("采用AI建议", "use_ai_suggestion"),
            ]:
                col = cols.get(canonical)
                if col:
                    parsed = self._parse_bool_value(row.get(col), default=None)
                    if parsed is not None:
                        state[state_key] = parsed

            for canonical, state_key in [
                ("AI建议名", "ai_suggested_stem"),
                ("AI建议原因", "ai_reason"),
                ("AI置信度", "ai_confidence"),
                ("AI审核结果", "ai_review"),
                ("AI审核风险", "ai_review_risk_level"),
                ("AI审核说明", "ai_review_comment"),
                ("状态", "status"),
                ("备注", "remark"),
            ]:
                col = cols.get(canonical)
                if col and not pd.isna(row.get(col)):
                    state[state_key] = str(row.get(col)).strip()

            new_state[original_name] = state

        missing_count = len(inventory_names - set(duplicate_names.tolist()))
        return new_state, {
            "ok": True,
            "template_rows": len(df),
            "parsed_rows": parsed_rows,
            "updated": updated,
            "skipped_empty": skipped_empty,
            "missing_count": missing_count,
            "unmatched": unmatched,
            "errors": [],
        }

    def _show_template_import_result(self, result, tone="success"):
        summary = f"""
        模板总行数：<b>{result.get('template_rows', 0)}</b><br>
        成功更新数量：<b>{result.get('updated', 0)}</b><br>
        空新文件名跳过数量：<b>{result.get('skipped_empty', 0)}</b><br>
        当前任务缺失但模板未覆盖数量：<b>{result.get('missing_count', 0)}</b><br>
        模板中未匹配原文件名数量：<b>{len(result.get('unmatched', []))}</b>
        """
        with self.template_result_output:
            clear_output(wait=True)
            display(HTML(panel_html("Excel 模板导入结果", summary, tone)))
            if result.get("unmatched"):
                display(HTML(dataframe_html(pd.DataFrame(result["unmatched"]), title="模板中未匹配原文件名明细", max_rows=80)))
            if result.get("errors"):
                display(HTML(panel_html("导入错误明细", "<br>".join(escape(e) for e in result["errors"]), "error")))

    def _on_export_excel_template_clicked(self, _):
        if not self.upload_completed:
            with self.template_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("无法导出 Excel 模板", "请先上传待改名文件。", "warning")))
            return
        try:
            template_path = self._build_excel_template()
            self._download_file(template_path)
            with self.template_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("Excel 模板已导出", f"已生成并触发下载：<b>{escape(template_path.name)}</b>", "success")))
            self._log("Excel 模板已导出。")
        except Exception as e:
            with self.template_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("Excel 模板导出失败", escape(e), "error")))
            self._log(f"Excel 模板导出失败：{e}")

    def _on_import_excel_template_clicked(self, _):
        if not self.upload_completed:
            with self.template_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("无法上传 Excel 模板", "请先上传待改名文件。", "warning")))
            return
        try:
            self._set_status("等待上传 Excel 模板", "info")
            uploaded = files.upload()
            if not uploaded:
                with self.template_result_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("未上传模板", "请选择 .xlsx 模板文件。", "warning")))
                return
            if len(uploaded) > 1:
                with self.template_result_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("模板上传失败", "一次只允许上传一个 .xlsx 模板文件。", "error")))
                return
            name, content = next(iter(uploaded.items()))
            if not str(name).lower().endswith(".xlsx"):
                with self.template_result_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("模板上传失败", "只接受 .xlsx 文件。", "error")))
                return
            template_upload_path = self.main_system_dir / Path(name).name
            self.main_system_dir.mkdir(parents=True, exist_ok=True)
            with open(template_upload_path, "wb") as f:
                f.write(content)
            temp_colab_path = Path("/content") / Path(name).name
            if temp_colab_path.exists() and temp_colab_path != template_upload_path:
                safe_remove(temp_colab_path)

            df = pd.read_excel(template_upload_path, sheet_name=0)
            new_state, result = self._import_template_df_to_state(df)
            if not result["ok"]:
                self.validation_passed = False
                self.latest_valid_plan_df = None
                self._show_template_import_result(result, tone="error")
                self._set_status("Excel 模板导入失败", "error")
                self._refresh_buttons()
                return

            self.rename_state = new_state
            self._refresh_online_table_from_state()
            self.has_template_imported = True
            self._set_stage("template_imported")
            self._mark_dirty()
            self._show_template_import_result(result, tone="success")
            self._set_status("Excel 模板已导入，等待重新校验", "warning")
            self._log(f"Excel 模板导入完成：更新 {result['updated']} 条。")
        except Exception as e:
            with self.template_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("Excel 模板读取或导入失败", escape(e), "error")))
            self.validation_passed = False
            self.latest_valid_plan_df = None
            self._refresh_buttons()
            self._log(f"Excel 模板导入失败：{e}")

    # --------------------------------------------------------
    # DSV4Flash optional AI suggestions
    # --------------------------------------------------------
    def _ai_scene_defaults(self):
        return {
            "通用清洗型": {
                "description": "去除乱码、多余空格、重复符号和 copy/final/临时/测试等无意义词，保留核心含义。",
                "format": "",
                "params": {
                    "是否保留原编号": True,
                    "是否保留日期": True,
                    "是否统一大小写": False,
                    "需要删除的词": ["copy", "final", "临时", "测试"],
                    "命名语言风格": "混合",
                },
            },
            "编号规则型": {
                "description": "识别文件名前缀编号，根据编号映射、历史记录或规则字典生成标准名称。",
                "format": "编号_标准名称",
                "params": {
                    "是否保留前缀编号": True,
                    "编号映射规则": DEFAULT_RULE_RENAME_MAP,
                    "未命中编号如何处理": "保留原主体并清理无意义后缀",
                    "需要移除的后缀": ["_vlm1_spd1x"],
                },
            },
            "电商数据文件型": {
                "description": "识别平台、数据类型、对象和日期，生成统一电商数据文件名。",
                "format": "平台_数据类型_对象_日期",
                "params": {
                    "平台": "京东 / 抖音 / 小红书 / 淘宝 / 其他",
                    "数据类型": "关键词 / 商品 / 店铺 / 交易 / 流量 / 自定义",
                    "日期或月份": "",
                    "品类 / 关键词 / 店铺 / 对象": "",
                    "平台缩写规则": {"京东": "JD", "抖音": "DY", "小红书": "XHS", "淘宝": "TB"},
                },
            },
            "图片素材整理型": {
                "description": "按产品名、用途和序号统一图片素材命名。",
                "format": "产品名_用途_序号",
                "params": {
                    "产品名": "",
                    "素材用途": "主图 / 详情图 / 场景图 / 证书 / 其他",
                    "起始序号": 1,
                    "序号位数": 2,
                    "是否保留原日期": True,
                },
            },
            "音频/视频素材型": {
                "description": "识别编号，清理素材后缀，统一音视频素材命名。",
                "format": "编号_角色_版本",
                "params": {
                    "是否保留前缀编号": True,
                    "是否移除 _vlm1_spd1x": True,
                    "语言标记": "",
                    "音色 / 角色": "",
                    "版本号": "",
                },
            },
            "文档归档型": {
                "description": "生成合同、说明书、报告、发票、扫描件等归档名称。",
                "format": "文档类型_主体_日期_版本",
                "params": {
                    "文档类型": "",
                    "主体名称": "",
                    "日期": "",
                    "版本号": "",
                    "是否保留原编号": True,
                },
            },
            "自定义规则型": {
                "description": "按用户自然语言规则生成建议名。",
                "format": "",
                "params": {
                    "用户自然语言规则": "",
                    "输出语言": "混合",
                    "是否严格保留原编号": True,
                    "是否严格保留原日期": True,
                },
            },
        }

    def _on_ai_scene_changed(self, _):
        scene = getattr(self, "ai_scene_dropdown", None).value if hasattr(self, "ai_scene_dropdown") else "通用清洗型"
        config = self._ai_scene_defaults().get(scene, self._ai_scene_defaults()["通用清洗型"])
        if hasattr(self, "ai_scene_description"):
            self.ai_scene_description.value = panel_html(scene, escape(config["description"]), "gray")
        if hasattr(self, "ai_naming_format"):
            self.ai_naming_format.value = config["format"]
        if hasattr(self, "ai_scene_params"):
            self.ai_scene_params.value = json.dumps(config["params"], ensure_ascii=False, indent=2)

    def _get_ai_client(self):
        if not self.ai_enabled_checkbox.value:
            raise RuntimeError("请先启用 AI。")
        api_key = self.ai_api_key_text.value.strip()
        if not api_key:
            raise RuntimeError("请先填写 API Key 并启用 AI。")
        base_url = self.ai_base_url_text.value.strip() or None
        model = self.ai_model_text.value.strip() or "dsv4flash"
        return DSV4FlashClient(api_key=api_key, model=model, base_url=base_url)

    def _parse_scene_params(self):
        text = self.ai_scene_params.value.strip()
        if not text:
            return {}
        try:
            return json.loads(text)
        except Exception as e:
            raise RuntimeError(f"场景参数必须是有效 JSON：{e}")

    def _build_ai_payload(self):
        self._read_online_table_to_state()
        current_files = []
        for item in self.file_inventory:
            state = self.rename_state.get(item["original_name"], {})
            current_files.append({
                "original_name": item["original_name"],
                "original_stem": item["stem"],
                "extension": item["extension"],
                "current_new_stem": state.get("new_stem", ""),
                "keep_extension": bool(state.get("keep_extension", True)),
                "existing_ai_suggested_stem": state.get("ai_suggested_stem", ""),
            })
        return {
            "scene": self.ai_scene_dropdown.value,
            "user_instruction": self.ai_user_instruction.value,
            "naming_format": self.ai_naming_format.value,
            "scene_params": self._parse_scene_params(),
            "files": current_files,
            "history_context": self._build_history_context_for_ai(current_files, limit=100),
            "rules_summary": [
                "只生成文件名主体，不要包含扩展名。",
                "不要修改扩展名。",
                "原文件名必须逐字匹配。",
                "不要遗漏文件。",
                "建议名不能包含路径。",
                "不要输出 Markdown。",
                "必须返回严格 JSON。",
            ],
        }

    def _strip_ai_extension(self, suggested_stem, item):
        value = "" if suggested_stem is None else str(suggested_stem).strip()
        extension = item.get("extension", "")
        if extension and value.lower().endswith(extension.lower()):
            value = value[:-len(extension)].rstrip(" .")
        return value

    def _apply_ai_response_to_state(self, ai_data):
        if not isinstance(ai_data, dict):
            raise RuntimeError("AI 返回结果不是 JSON 对象。")
        items = ai_data.get("items")
        if not isinstance(items, list):
            raise RuntimeError("AI 返回 JSON 缺少 items 数组。")

        inventory_by_name = {item["original_name"]: item for item in self.file_inventory}
        result_rows = []
        unmatched = []
        skipped_empty = 0
        updated = 0
        overwritten_existing = 0

        for item in items:
            if not isinstance(item, dict):
                continue
            original_name = str(item.get("original_name", "")).strip()
            if original_name not in inventory_by_name:
                unmatched.append({"original_name": original_name, "说明": "AI 返回的 original_name 无法匹配当前文件"})
                continue
            source_item = inventory_by_name[original_name]
            suggested = self._strip_ai_extension(item.get("suggested_stem", ""), source_item)
            if not suggested:
                skipped_empty += 1
                continue
            clean_suggested = sanitize_filename_component(suggested)
            state = self.rename_state.get(original_name, {})
            if state.get("ai_suggested_stem"):
                overwritten_existing += 1
            state["ai_suggested_stem"] = clean_suggested
            state["ai_reason"] = str(item.get("reason", "")).strip()
            confidence = str(item.get("confidence", "")).strip()
            state["ai_confidence"] = confidence if confidence in {"高", "中", "低"} else ""
            self.rename_state[original_name] = state
            updated += 1
            result_rows.append({
                "原文件名": original_name,
                "当前新文件名主体": state.get("new_stem", ""),
                "AI建议名": clean_suggested,
                "置信度": state.get("ai_confidence", ""),
                "AI建议原因": state.get("ai_reason", ""),
            })

        result_df = pd.DataFrame(result_rows)
        stats = {
            "input_count": len(self.file_inventory),
            "returned_count": len(items),
            "updated": updated,
            "skipped_empty": skipped_empty,
            "unmatched": len(unmatched),
            "overwritten_existing": overwritten_existing,
            "high": int((result_df["置信度"] == "高").sum()) if not result_df.empty else 0,
            "mid": int((result_df["置信度"] == "中").sum()) if not result_df.empty else 0,
            "low": int((result_df["置信度"] == "低").sum()) if not result_df.empty else 0,
        }
        return result_df, pd.DataFrame(unmatched), stats

    def _show_ai_result(self, result_df, unmatched_df, stats):
        summary = f"""
        输入文件数：<b>{stats['input_count']}</b><br>
        AI 返回条数：<b>{stats['returned_count']}</b><br>
        成功写入建议数：<b>{stats['updated']}</b><br>
        跳过空建议数：<b>{stats['skipped_empty']}</b><br>
        未匹配 original_name 数：<b>{stats['unmatched']}</b><br>
        高 / 中 / 低置信度数量：<b>{stats['high']} / {stats['mid']} / {stats['low']}</b><br>
        已覆盖原 AI建议名字段数量：<b>{stats['overwritten_existing']}</b>
        """
        with self.ai_result_output:
            clear_output(wait=True)
            display(HTML(panel_html("AI 建议生成结果", summary, "success" if stats["updated"] else "warning")))
            if not result_df.empty:
                display(HTML(dataframe_html(result_df, title="AI 建议明细表", max_rows=100)))
            if not unmatched_df.empty:
                display(HTML(dataframe_html(unmatched_df, title="AI 返回未匹配 original_name 明细", max_rows=80)))

    def _on_test_ai_connection_clicked(self, _):
        try:
            client = self._get_ai_client()
            result = client.test_connection()
            self.ai_available = True
            with self.ai_status_output:
                clear_output(wait=True)
                display(HTML(panel_html("AI 连接测试成功", escape(json.dumps(result, ensure_ascii=False)), "success")))
        except Exception as e:
            self.ai_available = False
            with self.ai_status_output:
                clear_output(wait=True)
                display(HTML(panel_html("AI 连接测试失败", escape(e), "error")))
            self._log(f"AI 连接测试失败：{e}")

    def _on_generate_ai_suggestions_clicked(self, _):
        if not self.upload_completed:
            return
        try:
            client = self._get_ai_client()
            payload = self._build_ai_payload()
            ai_data = client.suggest_names(payload)
            result_df, unmatched_df, stats = self._apply_ai_response_to_state(ai_data)
            self.latest_ai_suggestion_result = {
                "raw": ai_data,
                "stats": stats,
                "result_df": result_df,
                "unmatched_df": unmatched_df,
            }
            self.has_ai_suggestions = stats["updated"] > 0
            self._set_stage("ai_suggested")
            self._refresh_online_table_from_state()
            self._mark_dirty()
            self._show_ai_result(result_df, unmatched_df, stats)
            self._log(f"AI 建议生成完成：写入 {stats['updated']} 条。")
        except Exception as e:
            with self.ai_result_output:
                clear_output(wait=True)
                display(HTML(panel_html("AI 建议生成失败", escape(e), "error")))
            self._log(f"AI 建议生成失败：{e}")

    def _on_apply_selected_ai_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        applied = 0
        for state in self.rename_state.values():
            if state.get("selected") and state.get("ai_suggested_stem"):
                state["new_stem"] = state["ai_suggested_stem"]
                applied += 1
        self._refresh_online_table_from_state()
        self._mark_dirty()
        with self.ai_result_output:
            display(HTML(panel_html("采用选中 AI 建议", f"已采用 <b>{applied}</b> 条选中 AI 建议。", "success" if applied else "warning")))

    def _on_apply_all_ai_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        applied = 0
        for state in self.rename_state.values():
            if state.get("ai_suggested_stem"):
                state["new_stem"] = state["ai_suggested_stem"]
                applied += 1
        self._refresh_online_table_from_state()
        self._mark_dirty()
        with self.ai_result_output:
            display(HTML(panel_html("采用全部 AI 建议", f"已采用 <b>{applied}</b> 条 AI 建议。", "success" if applied else "warning")))

    def _on_clear_ai_suggestions_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        for state in self.rename_state.values():
            state["ai_suggested_stem"] = ""
            state["ai_reason"] = ""
            state["ai_confidence"] = ""
        self._refresh_online_table_from_state()
        self._mark_dirty()
        with self.ai_result_output:
            clear_output(wait=True)
            display(HTML(panel_html("AI 建议已清空", "已清空 AI建议名、AI建议原因和 AI置信度；当前新文件名主体未被修改。", "info")))

    def _local_rules_summary(self):
        return [
            "AI 只提供建议和解释，不能执行改名。",
            "最终执行前必须通过本地硬校验。",
            "本地代码负责文件名清洗、重复检测、路径安全、复制、报告和 ZIP。",
            "AI 软重复和审核风险只提示，不拦截。",
            "不发送文件内容，不发送 API Key，不发送绝对路径。",
        ]

    def _current_stage(self):
        if self.latest_download_path and Path(self.latest_download_path).exists():
            return "executed"
        if self.validation_passed:
            return "validation_passed"
        if self.latest_validation_result and not self.latest_validation_result.get("ok"):
            return "validation_failed"
        if any(s.get("ai_review") for s in self.rename_state.values()):
            return "ai_reviewed"
        if any(s.get("ai_suggested_stem") for s in self.rename_state.values()):
            return "ai_suggested"
        if self.upload_completed:
            return "uploaded"
        return "waiting_upload"

    def _fixed_guidance(self, stage=None):
        stage = stage or self._current_stage()
        guide_map = {
            "waiting_upload": ("请先上传需要改名的文件。系统只会处理文件名，不会修改文件内容。", ["点击“上传待改名文件”", "上传后检查在线表格"]),
            "uploaded": ("已生成在线编辑表格。可直接修改“新文件名（不含扩展名）”，也可使用批量粘贴或 Excel 模板导入。", ["填写或粘贴新文件名", "点击“校验当前命名”"]),
            "editing": ("当前处于在线编辑阶段。建议先完成新文件名填写，然后点击校验当前命名。", ["检查新文件名是否完整", "点击校验当前命名"]),
            "pasted": ("批量粘贴后需要重新校验。", ["检查表格是否同步更新", "点击校验当前命名"]),
            "history_suggested": ("历史推荐已写入 AI建议名列，但尚未应用到最终新文件名。", ["采用需要的历史推荐", "重新校验"]),
            "ai_suggested": ("AI 建议已写入 AI建议名列，但尚未应用到最终新文件名。请逐行、选中或全部采用建议后重新校验。", ["采用 AI 建议", "重新校验"]),
            "validation_failed": ("当前命名未通过本地校验。请根据错误明细修复后重新校验。", ["查看错误明细", "修复后重新校验"]),
            "validation_passed": ("当前命名已通过本地校验，可以执行批量改名。执行后会生成 ZIP 和 rename_report.xlsx。", ["点击执行批量改名", "下载结果 ZIP"]),
            "executed": ("执行已完成。可以下载 ZIP，也可以开启新任务；历史记录已保留。", ["检查 rename_report.xlsx", "需要时重新下载结果包"]),
            "template_imported": ("Excel 模板已导入，当前表格已更新。导入不会绕过本地校验。", ["检查在线表格", "重新校验"]),
            "rule_module_ready": ("规则字典模块可独立上传文件并按编号规则处理，不影响主流程。", ["设置 ZIP 名称", "上传并按规则处理"]),
            "ai_reviewed": ("AI 审核结果已写入表格。审核只提示风险，不替代本地校验。", ["查看 AI 审核说明", "必要时修改后重新校验"]),
        }
        return guide_map.get(stage, guide_map["waiting_upload"])

    def _validation_summary(self):
        result = self.latest_validation_result or {}
        if not result:
            return {}
        return {
            "ok": bool(result.get("ok")),
            "issue_count": len(result.get("issue_df", pd.DataFrame())),
            "collision_count": len(result.get("collision_df", pd.DataFrame())),
            "warning_count": len(result.get("warning_df", pd.DataFrame())),
        }

    def _detect_local_misoperation_warnings(self, plan_df=None, rename_state=None):
        rename_state = rename_state or self.rename_state
        if plan_df is None:
            try:
                plan_df = self._build_rename_plan()
            except Exception:
                plan_df = pd.DataFrame()
        warnings = []
        total = len(plan_df) if plan_df is not None else 0
        if total == 0:
            return warnings

        unchanged = int((~plan_df["是否实际改名"]).sum())
        empty_new = sum(1 for state in rename_state.values() if not str(state.get("new_stem", "")).strip())
        no_ext = int((~plan_df["是否保留原扩展名"]).sum())
        contains_ext = int(plan_df.get("__contained_original_extension", pd.Series(dtype=bool)).sum()) if "__contained_original_extension" in plan_df.columns else 0
        same_stem = plan_df["清洗后的新文件名主体"].astype(str).str.lower().value_counts()
        suspicious_uploads = [item["original_name"] for item in self.file_inventory if re.search(r"(renamed_files|rule_renamed_files|rename_report)", item["original_name"], re.I)]
        ai_unapplied = sum(1 for s in rename_state.values() if s.get("ai_suggested_stem") and s.get("new_stem") != s.get("ai_suggested_stem"))

        def add(t, level, count, detail):
            warnings.append({"type": t, "level": level, "count": int(count), "ratio": round(count / total, 4) if total else 0, "detail": detail})

        if unchanged / total > 0.8:
            add("大量文件未改名", "warning", unchanged, f"当前 {unchanged}/{total} 个文件最终名称与原文件名一致，可能尚未完成编辑。")
        if empty_new / total > 0.3:
            add("大量 new_stem 为空", "warning", empty_new, f"当前 {empty_new}/{total} 个文件新文件名主体为空。")
        if no_ext / total > 0.3:
            add("大量取消保留扩展名", "warning", no_ext, f"当前 {no_ext}/{total} 个文件将不保留原扩展名。")
        if contains_ext / total > 0.3:
            add("大量用户填写值包含原扩展名", "warning", contains_ext, f"当前 {contains_ext}/{total} 个填写值包含原扩展名，系统会自动去除重复扩展名。")
        near_dup_count = int(same_stem[same_stem > 1].sum())
        if near_dup_count:
            add("大量文件写成相同或近似名称", "warning", near_dup_count, "多个文件清洗后的新主体相同或近似，请确认是否符合预期。")
        if suspicious_uploads:
            add("疑似上传结果包或报告", "warning", len(suspicious_uploads), "上传文件名包含 renamed_files、rule_renamed_files 或 rename_report 特征，可能误把输出结果再次上传。")
        if total > 50 and unchanged == total:
            add("大量文件全部未修改", "warning", unchanged, "当前文件数量较大且全部未改名，可能尚未开始编辑。")
        if ai_unapplied:
            add("AI 或历史建议未采用", "info", ai_unapplied, "已有 AI建议名，但尚未采用到新文件名主体。")
        return warnings

    def _review_payload(self):
        self._read_online_table_to_state()
        plan_df = self._build_rename_plan()
        current_files = []
        for _, row in plan_df.iterrows():
            current_files.append({
                "original_name": row["原文件名"],
                "original_stem": row["原文件名主体"],
                "extension": row["原扩展名"],
                "current_new_stem": row["用户填写的新文件名主体"],
                "keep_extension": bool(row["是否保留原扩展名"]),
                "final_name_preview": row["最终输出文件名"],
                "ai_suggested_stem": row["AI建议名"],
                "ai_reason": row["AI建议原因"],
            })
        return {
            "current_files": current_files,
            "rename_plan_preview": self._plan_display_df(plan_df, max_rows=200).to_dict(orient="records"),
            "validation_summary": self._validation_summary(),
            "warning_summary": self._detect_local_misoperation_warnings(plan_df, self.rename_state),
            "history_context": self._build_history_context_for_ai(current_files, limit=100),
            "rules_summary": self._local_rules_summary(),
            "user_instruction": self.ai_review_instruction.value,
        }

    def _call_ai_json(self, system_prompt, payload, timeout=60):
        client = self._get_ai_client()
        return client.chat_json([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
        ], timeout=timeout)

    def _on_ai_review_names_clicked(self, _):
        if not self.upload_completed:
            return
        try:
            payload = self._review_payload()
            system_prompt = (
                "你是文件批量改名审核助手。只根据文件名、当前新文件名、历史记录和工具规则进行审核。"
                "不读取、不推测文件内容。本地硬重复由代码处理，AI 只判断语义风险和流程风险。"
                "输出严格 JSON，不得输出 Markdown。original_name 必须与输入逐字匹配。"
                "review_status 只能为：通过、建议修改、高风险。risk_level 只能为：低、中、高。comment 简短可执行。"
            )
            data = self._call_ai_json(system_prompt, payload)
            self.latest_ai_review_result = data
            items = data.get("items")
            if not isinstance(items, list):
                raise RuntimeError("AI 审核返回 JSON 缺少 items 数组。")
            names = set(self.rename_state.keys())
            rows = []
            unmatched = []
            for item in items:
                original_name = str(item.get("original_name", "")).strip()
                if original_name not in names:
                    unmatched.append({"original_name": original_name, "说明": "无法匹配当前文件"})
                    continue
                status = str(item.get("review_status", "")).strip()
                risk = str(item.get("risk_level", "")).strip()
                comment = str(item.get("comment", "")).strip()
                if status not in {"通过", "建议修改", "高风险"}:
                    status = "建议修改"
                if risk not in {"低", "中", "高"}:
                    risk = "中"
                self.rename_state[original_name]["ai_review"] = status
                self.rename_state[original_name]["ai_review_risk_level"] = risk
                self.rename_state[original_name]["ai_review_comment"] = comment
                rows.append({"原文件名": original_name, "AI审核结果": status, "AI审核风险": risk, "AI审核说明": comment})
            self._refresh_online_table_from_state()
            self._mark_dirty()
            with self.ai_review_output:
                clear_output(wait=True)
                summary = data.get("summary", {})
                display(HTML(panel_html("AI 审核完成", escape(json.dumps(summary, ensure_ascii=False)), "success")))
                if rows:
                    display(HTML(dataframe_html(pd.DataFrame(rows), title="AI 审核明细", max_rows=100)))
                if data.get("semantic_duplicate_risks"):
                    display(HTML(dataframe_html(pd.DataFrame(data["semantic_duplicate_risks"]), title="AI 软重复风险（仅提示，不拦截）", max_rows=80)))
                if data.get("operation_risks"):
                    display(HTML(dataframe_html(pd.DataFrame(data["operation_risks"]), title="AI 操作风险提示", max_rows=80)))
                if unmatched:
                    display(HTML(dataframe_html(pd.DataFrame(unmatched), title="AI 审核未匹配文件", max_rows=80)))
        except Exception as e:
            with self.ai_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("AI 审核失败", escape(e), "error")))
            self._log(f"AI 审核失败：{e}")

    def _on_clear_ai_review_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        for state in self.rename_state.values():
            state["ai_review"] = ""
            state["ai_review_risk_level"] = ""
            state["ai_review_comment"] = ""
        self._refresh_online_table_from_state()
        self._mark_dirty()
        with self.ai_review_output:
            clear_output(wait=True)
            display(HTML(panel_html("AI 审核已清空", "已清空 AI审核结果、AI审核风险和 AI审核说明；未修改新文件名和 AI建议名。", "info")))

    def _on_ai_explain_issues_clicked(self, _):
        try:
            payload = {
                "validation_summary": self._validation_summary(),
                "issues": (self.latest_validation_result or {}).get("issue_df", pd.DataFrame()).to_dict(orient="records") if self.latest_validation_result else [],
                "collisions": (self.latest_validation_result or {}).get("collision_df", pd.DataFrame()).to_dict(orient="records") if self.latest_validation_result else [],
                "warnings": (self.latest_validation_result or {}).get("warning_df", pd.DataFrame()).to_dict(orient="records") if self.latest_validation_result else [],
                "local_misoperation_warnings": self._detect_local_misoperation_warnings(),
                "rules_summary": self._local_rules_summary(),
            }
            system_prompt = (
                "你是批量改名问题解释助手。根据本地校验错误和误操作预警给出简短修复建议。"
                "不得声称已修改文件，不得要求跳过本地校验。必须返回严格 JSON："
                '{"summary":"...","fix_steps":["..."],"risk_notes":["..."],"next_action":"..."}'
            )
            data = self._call_ai_json(system_prompt, payload)
        except Exception:
            data = {
                "summary": "当前问题需要先根据本地校验和预警明细处理。",
                "fix_steps": ["查看错误表和冲突表", "修改在线表格或重新导入模板", "重新点击校验当前命名"],
                "risk_notes": ["AI 解释不可用时，本地校验结果仍是执行前的依据。"],
                "next_action": "请先修复本地校验错误，再重新校验。",
            }
        with self.ai_review_output:
            display(HTML(panel_html(
                "AI 当前问题解释",
                escape(data.get("summary", "")) + "<br><b>下一步：</b>" + escape(data.get("next_action", "")),
                "info",
            )))
            if data.get("fix_steps"):
                display(HTML(dataframe_html(pd.DataFrame({"修复步骤": data["fix_steps"]}), title="建议修复步骤", max_rows=20)))
            if data.get("risk_notes"):
                display(HTML(dataframe_html(pd.DataFrame({"风险提示": data["risk_notes"]}), title="风险提示", max_rows=20)))

    def _guidance_payload(self):
        stage = self._current_stage()
        return {
            "app_stage": stage,
            "uploaded_file_count": len(self.file_inventory),
            "has_ai_suggestions": any(s.get("ai_suggested_stem") for s in self.rename_state.values()),
            "has_history_suggestions": any("历史" in str(s.get("ai_reason", "")) for s in self.rename_state.values()),
            "latest_validation_summary": self._validation_summary(),
            "latest_warning_summary": self._detect_local_misoperation_warnings(),
            "rules_summary": self._local_rules_summary(),
        }

    def _on_refresh_ai_guidance_clicked(self, _):
        stage = self._current_stage()
        try:
            if self.ai_enabled_checkbox.value and self.ai_api_key_text.value.strip():
                system_prompt = (
                    "你是批量改名流程指引助手。不要声称已修改文件，不要要求跳过校验，不要要求上传文件内容。"
                    '必须返回严格 JSON：{"guidance":"...","next_steps":["..."]}'
                )
                data = self._call_ai_json(system_prompt, self._guidance_payload(), timeout=45)
                guidance = data.get("guidance", "")
                steps = data.get("next_steps", [])
            else:
                guidance, steps = self._fixed_guidance(stage)
        except Exception:
            guidance, steps = self._fixed_guidance(stage)
        with self.ai_guidance_output:
            clear_output(wait=True)
            display(HTML(panel_html("当前步骤指引", escape(guidance), "info")))
            if steps:
                display(HTML(dataframe_html(pd.DataFrame({"下一步": steps}), title="建议下一步", max_rows=20)))

    def _qa_context(self, question):
        return {
            "app_stage": self._current_stage(),
            "uploaded_file_count": len(self.file_inventory),
            "has_ai_suggestions": any(s.get("ai_suggested_stem") for s in self.rename_state.values()),
            "has_history_suggestions": any("历史" in str(s.get("ai_reason", "")) for s in self.rename_state.values()),
            "latest_validation_summary": self._validation_summary(),
            "local_warnings": self._detect_local_misoperation_warnings(),
            "rules_summary": self._local_rules_summary(),
            "question": question,
        }

    def _fallback_qa_answer(self, question):
        if "不能执行" in question:
            return {"answer": "执行按钮通常在尚未通过本地校验，或表格在校验后又被修改时不可用。", "suggested_action": "请点击“校验当前命名”，通过后再执行批量改名。"}
        if "重复" in question:
            return {"answer": "重复文件名会被本地硬校验拦截，尤其是大小写不敏感的重复。", "suggested_action": "修改重复文件名中的至少一个，然后重新校验。"}
        if "AI 建议" in question:
            return {"answer": "AI 建议只写入 AI建议名列，不会自动覆盖新文件名。", "suggested_action": "确认建议后点击采用选中或采用全部 AI 建议，再重新校验。"}
        return {"answer": "当前工具按上传、编辑、校验、执行、下载的顺序工作，AI 只做辅助建议和解释。", "suggested_action": "先查看当前状态区，再按提示完成下一步。"}

    def _on_ai_qa_ask_clicked(self, _):
        question = self.ai_qa_question.value.strip()
        if not question:
            with self.ai_qa_output:
                clear_output(wait=True)
                display(HTML(panel_html("问题为空", "请输入问题，或点击快捷问题。", "warning")))
            return
        try:
            if self.ai_enabled_checkbox.value and self.ai_api_key_text.value.strip():
                system_prompt = (
                    "你是批量改名工具的新用户问答助手。回答简短明确，不得编造不存在的按钮，"
                    '不得声称已执行操作，不得要求跳过校验。必须返回严格 JSON：{"answer":"...","suggested_action":"..."}'
                )
                data = self._call_ai_json(system_prompt, self._qa_context(question), timeout=45)
            else:
                data = self._fallback_qa_answer(question)
        except Exception:
            data = self._fallback_qa_answer(question)
        with self.ai_qa_output:
            clear_output(wait=True)
            display(HTML(panel_html("问答助手", f"{escape(data.get('answer', ''))}<br><b>建议操作：</b>{escape(data.get('suggested_action', ''))}", "info")))

    def _on_quick_question_clicked(self, question):
        self.ai_qa_question.value = question
        self._on_ai_qa_ask_clicked(None)

    def _on_refresh_history_clicked(self, _):
        self._show_history_review(mode="recent")
        self._set_status("历史记录已刷新", "info")

    def _on_show_recent_history_clicked(self, _):
        self._show_history_review(mode="recent")

    def _on_show_all_history_clicked(self, _):
        self._show_history_review(mode="all")

    def _on_generate_history_copy_clicked(self, _):
        try:
            self.history_copy_text.value = self._generate_history_copy_text(limit=30)
            self._set_status("精简复制文本已生成", "info")
        except Exception as e:
            with self.history_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("精简复制文本生成失败", escape(e), "error")))
            self._log(f"精简复制文本生成失败：{e}")

    def _on_export_history_clicked(self, _):
        try:
            df = self._load_history_df()
            if df.empty:
                with self.history_review_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("暂无历史记录可导出", "成功执行批量改名后再导出历史记录。", "warning")))
                return
            if not self.history_xlsx_path.exists():
                self._save_history_df(df)
            self._download_file(self.history_xlsx_path)
            self._set_status("历史记录下载已触发", "success")
            with self.history_review_output:
                display(HTML(panel_html("导出历史记录", "已触发下载 rename_history.xlsx。", "success")))
        except Exception as e:
            with self.history_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("历史记录导出失败", escape(e), "error")))
            self._log(f"历史记录导出失败：{e}")

    def _on_clear_history_clicked(self, _):
        try:
            safe_remove(self.history_xlsx_path)
            safe_remove(self.history_csv_path)
            self.history_dir.mkdir(parents=True, exist_ok=True)
            self.last_history_df = pd.DataFrame(columns=HISTORY_COLUMNS)
            self.history_copy_text.value = ""
            with self.history_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("历史记录已清空", "rename_history.xlsx 和 rename_history.csv 已清空；当前任务和 rename_state 不受影响。", "success")))
            self._set_status("历史记录已清空", "success")
            self._log("历史记录已清空。")
        except Exception as e:
            with self.history_review_output:
                clear_output(wait=True)
                display(HTML(panel_html("清空历史记录失败", escape(e), "error")))
            self._log(f"清空历史记录失败：{e}")

    def _on_generate_history_suggestions_clicked(self, _):
        if not self.upload_completed:
            return
        try:
            suggestion_df = self._generate_history_based_suggestions()
            self.last_suggestion_df = suggestion_df
            self.has_history_suggestions = not suggestion_df.empty
            self._refresh_online_table_from_state()
            self._set_stage("history_suggested")
            self._mark_dirty()
            self._show_history_suggestion_result(suggestion_df)
            self._log(f"历史推荐生成完成：{len(suggestion_df)} 条。")
        except Exception as e:
            with self.history_suggestion_output:
                clear_output(wait=True)
                display(HTML(panel_html("历史推荐生成失败", escape(e), "error")))
            self._log(f"历史推荐生成失败：{e}")

    def _on_apply_selected_history_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        applied = 0
        for original_name, state in self.rename_state.items():
            if state.get("selected") and state.get("ai_suggested_stem"):
                state["new_stem"] = state["ai_suggested_stem"]
                applied += 1
            self._refresh_online_table_from_state()
            self._set_stage("ai_suggested")
            self._mark_dirty()
        with self.history_suggestion_output:
            display(HTML(panel_html("采用选中历史推荐", f"已采用 <b>{applied}</b> 条选中历史推荐。", "success" if applied else "warning")))
        self._log(f"采用选中历史推荐：{applied} 条。")

    def _on_apply_all_history_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        applied = 0
        for state in self.rename_state.values():
            if state.get("ai_suggested_stem"):
                state["new_stem"] = state["ai_suggested_stem"]
                applied += 1
            self._refresh_online_table_from_state()
            self._set_stage("ai_reviewed")
            self._mark_dirty()
        with self.history_suggestion_output:
            display(HTML(panel_html("采用全部历史推荐", f"已采用 <b>{applied}</b> 条历史推荐。", "success" if applied else "warning")))
        self._log(f"采用全部历史推荐：{applied} 条。")

    def _on_clear_history_suggestions_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        for state in self.rename_state.values():
            state["ai_suggested_stem"] = ""
            state["ai_reason"] = ""
            state["ai_confidence"] = ""
        self.last_suggestion_df = pd.DataFrame()
        self._refresh_online_table_from_state()
        self._mark_dirty()
        with self.history_suggestion_output:
            clear_output(wait=True)
            display(HTML(panel_html("历史推荐已清空", "已清空 AI建议名、AI建议原因和推荐置信度；用户当前新文件名主体未被修改。", "info")))
        self._log("历史推荐已清空。")

    # --------------------------------------------------------
    # History records
    # --------------------------------------------------------
    def _ensure_history_columns(self, df):
        if df is None or df.empty:
            return pd.DataFrame(columns=HISTORY_COLUMNS)
        df = df.copy()
        for col in HISTORY_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[HISTORY_COLUMNS]

    def _load_history_df(self, show_error=True):
        try:
            if self.history_xlsx_path.exists():
                df = pd.read_excel(self.history_xlsx_path)
            elif self.history_csv_path.exists():
                df = pd.read_csv(self.history_csv_path, encoding="utf-8-sig")
            else:
                df = pd.DataFrame(columns=HISTORY_COLUMNS)
            df = self._ensure_history_columns(df)
            self.last_history_df = df
            return df
        except Exception as e:
            if show_error:
                with self.history_review_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("历史记录读取失败", escape(e), "error")))
            self._log(f"历史记录读取失败：{e}")
            return pd.DataFrame(columns=HISTORY_COLUMNS)

    def _save_history_df(self, df):
        self.history_dir.mkdir(parents=True, exist_ok=True)
        df = self._ensure_history_columns(df)
        df.to_csv(self.history_csv_path, index=False, encoding="utf-8-sig")
        df.to_excel(self.history_xlsx_path, index=False)
        beautify_excel(self.history_xlsx_path)

    def _history_rows_from_plan(self, plan_df, task_id, timestamp):
        rows = []
        for _, row in plan_df.iterrows():
            original_stem = row["原文件名主体"]
            final_name = row["最终输出文件名"]
            final_stem, _, _ = split_filename(final_name)
            original_num = extract_prefix_number(original_stem)
            final_num = extract_prefix_number(final_stem)
            rows.append({
                "时间戳": timestamp,
                "任务批次ID": task_id,
                "原文件名": row["原文件名"],
                "原文件名主体": original_stem,
                "原扩展名": row["原扩展名"],
                "用户填写的新文件名主体": row["用户填写的新文件名主体"],
                "清洗后的新文件名主体": row["清洗后的新文件名主体"],
                "最终输出文件名": final_name,
                "是否保留原扩展名": row["是否保留原扩展名"],
                "是否实际改名": row["是否实际改名"],
                "处理说明": row["处理说明"],
                "AI建议名": row["AI建议名"],
                "AI建议原因": row["AI建议原因"],
                "AI置信度": row.get("AI置信度", ""),
                "AI审核结果": row["AI审核结果"],
                "AI审核风险": row.get("AI审核风险", ""),
                "AI审核说明": row.get("AI审核说明", ""),
                "状态": row["状态"],
                "original_stem_norm": normalize_for_history(original_stem),
                "final_stem_norm": normalize_for_history(final_stem),
                "original_prefix_number": original_num,
                "final_prefix_number": final_num,
                "original_tokens": tokenize_for_history(original_stem),
                "final_tokens": tokenize_for_history(final_stem),
                "pattern_key": f"prefix_number_{original_num}" if original_num else "",
            })
        return rows

    def _append_history_from_plan(self, plan_df):
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        task_id = time.strftime("task_%Y%m%d_%H%M%S")
        history_df = self._load_history_df(show_error=False)

        if not history_df.empty and "任务批次ID" in history_df.columns:
            while task_id in set(history_df["任务批次ID"].astype(str)):
                task_id = f"task_{time.strftime('%Y%m%d_%H%M%S')}_{int(time.time() * 1000) % 1000}"

        new_df = pd.DataFrame(self._history_rows_from_plan(plan_df, task_id, timestamp))
        merged_df = pd.concat([history_df, new_df], ignore_index=True)
        self._save_history_df(merged_df)
        self.last_history_df = merged_df
        return {"task_id": task_id, "count": len(new_df), "timestamp": timestamp}

    def _history_summary_html(self, df):
        total = len(df)
        batches = df["任务批次ID"].nunique() if total and "任务批次ID" in df.columns else 0
        latest = ""
        if total and "时间戳" in df.columns:
            latest = str(df["时间戳"].dropna().astype(str).iloc[-1])
        latest = latest or "暂无"
        return f"""
        历史总记录数：<b>{total}</b><br>
        历史任务批次数：<b>{batches}</b><br>
        最近一次改名时间：<b>{escape(latest)}</b>
        """

    def _history_preview_df(self, df, mode="recent"):
        if df.empty:
            return pd.DataFrame()
        cols = ["时间戳", "任务批次ID", "原文件名", "原文件名主体", "最终输出文件名", "清洗后的新文件名主体", "处理说明"]
        view_df = df[cols].copy()
        if mode == "all":
            return view_df.head(300)
        return view_df.tail(30).iloc[::-1]

    def _show_history_review(self, mode="recent"):
        df = self._load_history_df()
        with self.history_review_output:
            clear_output(wait=True)
            display(HTML(panel_html("历史改名记录回顾", self._history_summary_html(df), "info")))
            if df.empty:
                display(HTML(panel_html("暂无历史记录", "成功执行批量改名后，系统会自动追加历史记录。", "gray")))
            else:
                if mode == "all" and len(df) > 300:
                    display(HTML(panel_html("展示限制", "历史记录超过 300 条，页面仅展示前 300 条；可点击导出历史记录获取完整文件。", "warning")))
                title = "全部历史记录预览" if mode == "all" else "最近 30 条历史记录"
                display(HTML(dataframe_html(self._history_preview_df(df, mode=mode), title=title, max_rows=300)))

    def _generate_history_copy_text(self, limit=300):
        df = self._load_history_df()
        if df.empty:
            return "暂无历史记录可复制。"
        view_df = df.tail(min(limit, 300))
        fmt = self.history_copy_format.value
        lines = []
        if fmt == "arrow":
            lines = [f"{row['原文件名']} -> {row['最终输出文件名']}" for _, row in view_df.iterrows()]
        elif fmt == "stem":
            lines = ["原主体\t新主体"]
            lines.extend(f"{row['原文件名主体']}\t{row['清洗后的新文件名主体']}" for _, row in view_df.iterrows())
        else:
            lines = ["原文件名\t最终输出文件名"]
            lines.extend(f"{row['原文件名']}\t{row['最终输出文件名']}" for _, row in view_df.iterrows())
        return "\n".join(str(line) for line in lines)

    # --------------------------------------------------------
    # History based suggestions
    # --------------------------------------------------------
    def _pick_common_recent_final_stem(self, df):
        if df.empty:
            return "", "", 0
        work_df = df.copy()
        work_df["清洗后的新文件名主体"] = work_df["清洗后的新文件名主体"].astype(str)
        counts = work_df.groupby("清洗后的新文件名主体").size().reset_index(name="count")
        counts = counts.sort_values("count", ascending=False)
        top_count = int(counts.iloc[0]["count"])
        candidates = set(counts[counts["count"] == top_count]["清洗后的新文件名主体"].astype(str))
        recent = work_df[work_df["清洗后的新文件名主体"].astype(str).isin(candidates)].iloc[-1]
        return str(recent["清洗后的新文件名主体"]), str(recent.get("最终输出文件名", "")), top_count

    def _suggest_for_item_from_history(self, item, history_df):
        if history_df.empty:
            return "", "", ""

        original_name = item["original_name"]
        stem = item["stem"]
        stem_norm = normalize_for_history(stem)
        prefix_number = extract_prefix_number(stem)
        tokens = set(tokenize_for_history(stem).split())

        exact_df = history_df[history_df["原文件名"].astype(str) == original_name]
        suggestion, _, count = self._pick_common_recent_final_stem(exact_df)
        if suggestion:
            return suggestion, "高", f"命中历史完全匹配（出现 {count} 次）"

        stem_df = history_df[history_df["原文件名主体"].astype(str) == stem]
        suggestion, _, count = self._pick_common_recent_final_stem(stem_df)
        if suggestion:
            return suggestion, "高", f"命中历史主体匹配（出现 {count} 次）"

        norm_df = history_df[history_df["original_stem_norm"].astype(str) == stem_norm]
        suggestion, _, count = self._pick_common_recent_final_stem(norm_df)
        if suggestion:
            return suggestion, "高", f"命中历史主体规范化匹配（出现 {count} 次）"

        if prefix_number:
            prefix_df = history_df[history_df["original_prefix_number"].astype(str) == prefix_number]
            suggestion, _, count = self._pick_common_recent_final_stem(prefix_df)
            if suggestion and count >= 2:
                confidence = "高" if count >= 3 else "中"
                return suggestion, confidence, f"命中历史编号模式：{prefix_number}（出现 {count} 次）"
            if suggestion:
                return suggestion, "中", f"命中历史编号模式：{prefix_number}"

            similar_rows = []
            for _, row in history_df.iterrows():
                row_tokens = set(str(row.get("original_tokens", "")).split())
                if prefix_number and str(row.get("original_prefix_number", "")) == prefix_number:
                    overlap = len(tokens & row_tokens)
                    if overlap:
                        similar_rows.append(row)
            if similar_rows:
                similar_df = pd.DataFrame(similar_rows)
                suggestion, _, count = self._pick_common_recent_final_stem(similar_df)
                if suggestion:
                    return suggestion, "低", f"命中历史模式相似：编号 {prefix_number} 且存在相同命名片段"

        return "", "", ""

    def _generate_history_based_suggestions(self):
        self._read_online_table_to_state()
        history_df = self._load_history_df()
        rows = []
        if history_df.empty:
            for state in self.rename_state.values():
                state["ai_suggested_stem"] = ""
                state["ai_reason"] = ""
                state["ai_confidence"] = ""
            return pd.DataFrame(columns=["原文件名", "当前新文件名主体", "历史推荐名", "置信度", "推荐原因"])

        for item in self.file_inventory:
            original_name = item["original_name"]
            state = self.rename_state.get(original_name, {})
            suggestion, confidence, reason = self._suggest_for_item_from_history(item, history_df)
            state["ai_suggested_stem"] = suggestion
            state["ai_reason"] = reason
            state["ai_confidence"] = confidence
            self.rename_state[original_name] = state
            if suggestion:
                rows.append({
                    "原文件名": original_name,
                    "当前新文件名主体": state.get("new_stem", ""),
                    "历史推荐名": suggestion,
                    "置信度": confidence,
                    "推荐原因": reason,
                })
        return pd.DataFrame(rows)

    def _build_history_context_for_ai(self, current_files, limit=100):
        history_df = self._load_history_df(show_error=False)
        if history_df.empty:
            return {"recent_records": [], "frequent_patterns": []}
        recent_df = history_df.tail(limit)
        recent_records = []
        for _, row in recent_df.iterrows():
            recent_records.append({
                "original_name": row.get("原文件名", ""),
                "original_stem": row.get("原文件名主体", ""),
                "final_name": row.get("最终输出文件名", ""),
                "final_stem": row.get("清洗后的新文件名主体", ""),
                "timestamp": row.get("时间戳", ""),
            })
        frequent_patterns = []
        pattern_df = history_df[history_df["pattern_key"].astype(str) != ""]
        for pattern_key, group in pattern_df.groupby("pattern_key"):
            examples = []
            for _, row in group.tail(5).iterrows():
                examples.append({
                    "original_stem": row.get("原文件名主体", ""),
                    "final_stem": row.get("清洗后的新文件名主体", ""),
                })
            frequent_patterns.append({"pattern_key": pattern_key, "examples": examples})
            if len(frequent_patterns) >= 20:
                break
        return {
            "recent_records": recent_records,
            "frequent_patterns": frequent_patterns,
        }

    def _show_history_suggestion_result(self, suggestion_df):
        history_df = self._load_history_df(show_error=False)
        total_history = len(history_df)
        total = len(suggestion_df)
        high = int((suggestion_df["置信度"] == "高").sum()) if not suggestion_df.empty else 0
        mid = int((suggestion_df["置信度"] == "中").sum()) if not suggestion_df.empty else 0
        low = int((suggestion_df["置信度"] == "低").sum()) if not suggestion_df.empty else 0
        summary = f"""
        历史记录总数：<b>{total_history}</b><br>
        本次生成推荐数量：<b>{total}</b><br>
        高置信度数量：<b>{high}</b><br>
        中置信度数量：<b>{mid}</b><br>
        低置信度数量：<b>{low}</b>
        """
        with self.history_suggestion_output:
            clear_output(wait=True)
            display(HTML(panel_html("历史推荐结果", summary, "success" if total else "warning")))
            if not suggestion_df.empty:
                display(HTML(dataframe_html(suggestion_df, title="推荐明细表", max_rows=80)))

    # --------------------------------------------------------
    # Validation and rename plan
    # --------------------------------------------------------
    def _is_inside_dir(self, child: Path, parent: Path) -> bool:
        child_resolved = Path(child).resolve(strict=False)
        parent_resolved = Path(parent).resolve(strict=False)
        try:
            child_resolved.relative_to(parent_resolved)
            return True
        except ValueError:
            return False

    def _make_final_name(self, item, state):
        raw_new_stem = "" if state.get("new_stem") is None else str(state.get("new_stem"))
        keep_extension = bool(state.get("keep_extension", True))
        default_used = not raw_new_stem.strip()
        working_stem = item["stem"] if default_used else raw_new_stem
        contained_original_extension = False

        extension = item.get("extension", "")
        if extension and working_stem.strip().lower().endswith(extension.lower()):
            contained_original_extension = True
            working_stem = working_stem.strip()[:-len(extension)].rstrip(" .")

        clean_stem = sanitize_filename_component(working_stem)
        final_name = clean_stem + (extension if keep_extension and extension else "")
        final_name = Path(final_name).name

        notes = []
        if default_used:
            notes.append("新文件名主体为空，已默认使用原文件名主体")
        if contained_original_extension:
            notes.append("已去除用户填写中重复的原扩展名")
        if raw_new_stem.strip() and clean_stem != raw_new_stem.strip():
            notes.append("新文件名主体已按安全规则清洗")
        if not notes:
            notes.append("正常")

        return raw_new_stem, clean_stem, keep_extension, final_name, contained_original_extension, "；".join(notes)

    def _build_rename_plan(self):
        self._read_online_table_to_state()
        rows = []

        for item in self.file_inventory:
            original_name = item["original_name"]
            state = self.rename_state.get(original_name, {})
            raw_new_stem, clean_stem, keep_extension, final_name, contained_ext, note = self._make_final_name(item, state)
            src_path = Path(item["source_path"])
            dst_path = self.main_output_dir / final_name
            actual_renamed = original_name != final_name

            rows.append({
                "原文件名": original_name,
                "原文件名主体": item["stem"],
                "原扩展名": item["extension"],
                "用户填写的新文件名主体": raw_new_stem,
                "清洗后的新文件名主体": clean_stem,
                "是否保留原扩展名": keep_extension,
                "最终输出文件名": final_name,
                "是否实际改名": actual_renamed,
                "源文件路径": str(src_path),
                "目标文件路径": str(dst_path),
                "AI建议名": state.get("ai_suggested_stem", ""),
                "AI建议原因": state.get("ai_reason", ""),
                "AI置信度": state.get("ai_confidence", ""),
                "AI审核结果": state.get("ai_review", ""),
                "AI审核风险": state.get("ai_review_risk_level", ""),
                "AI审核说明": state.get("ai_review_comment", ""),
                "处理说明": note,
                "状态": "待执行",
                "__final_key": final_name.lower(),
                "__src_path": str(src_path),
                "__dst_path": str(dst_path),
                "__default_used": bool(not raw_new_stem.strip()),
                "__contained_original_extension": bool(contained_ext),
            })

        return pd.DataFrame(rows)

    def _validate_current_state(self):
        issue_rows = []
        warning_rows = []
        collision_rows = []

        if not self.file_inventory:
            issue_rows.append({"问题类型": "文件索引为空", "原文件名": "", "最终输出文件名": "", "说明": "请先上传待改名文件。"})
        if not self.rename_state:
            issue_rows.append({"问题类型": "编辑状态为空", "原文件名": "", "最终输出文件名": "", "说明": "rename_state 为空，请重新上传或重置任务。"})

        plan_df = self._build_rename_plan() if self.file_inventory and self.rename_state else pd.DataFrame()
        output_root = self.main_output_dir.resolve(strict=False)

        for _, row in plan_df.iterrows():
            original_name = row["原文件名"]
            final_name = str(row["最终输出文件名"])
            src_path = Path(row["__src_path"])
            dst_path = Path(row["__dst_path"])

            if not src_path.exists():
                issue_rows.append({"问题类型": "源文件不存在", "原文件名": original_name, "最终输出文件名": final_name, "说明": str(src_path)})
            if not final_name.strip():
                issue_rows.append({"问题类型": "最终输出文件名为空", "原文件名": original_name, "最终输出文件名": final_name, "说明": "请填写新文件名主体。"})
            if re.search(r"[\\/]", final_name) or Path(final_name).name != final_name:
                issue_rows.append({"问题类型": "文件名包含路径分隔符", "原文件名": original_name, "最终输出文件名": final_name, "说明": "最终输出文件名不允许包含目录路径。"})
            if final_name in {".", ".."} or ".." in Path(final_name).parts:
                issue_rows.append({"问题类型": "路径穿越风险", "原文件名": original_name, "最终输出文件名": final_name, "说明": "最终输出文件名存在路径穿越风险。"})
            if final_name.lower() == "rename_report.xlsx":
                issue_rows.append({"问题类型": "输出文件名占用报告文件名", "原文件名": original_name, "最终输出文件名": final_name, "说明": "rename_report.xlsx 会作为系统报告文件生成，请换用其他输出文件名。"})
            if not str(row["清洗后的新文件名主体"]).strip():
                issue_rows.append({"问题类型": "清洗后的文件名主体为空", "原文件名": original_name, "最终输出文件名": final_name, "说明": "请修改新文件名主体。"})
            if not self._is_inside_dir(dst_path, output_root):
                issue_rows.append({"问题类型": "目标路径越界", "原文件名": original_name, "最终输出文件名": final_name, "说明": str(dst_path)})

        if not plan_df.empty:
            duplicate_df = plan_df[plan_df["__final_key"].duplicated(keep=False)].copy()
            if not duplicate_df.empty:
                for _, row in duplicate_df.sort_values("__final_key").iterrows():
                    collision_rows.append({
                        "冲突文件名": row["最终输出文件名"],
                        "冲突Key": row["__final_key"],
                        "原文件名": row["原文件名"],
                        "说明": "最终输出文件名按大小写不敏感规则重复，必须修复后才能执行。",
                    })

            total = len(plan_df)
            unchanged_count = int((~plan_df["是否实际改名"]).sum())
            default_count = int(plan_df["__default_used"].sum())
            no_extension_count = int((~plan_df["是否保留原扩展名"]).sum())
            contained_ext_count = int(plan_df["__contained_original_extension"].sum())
            threshold = max(3, int(total * 0.5)) if total else 3

            if unchanged_count >= threshold:
                warning_rows.append({"提醒类型": "大量文件未改名", "数量": unchanged_count, "说明": "较多最终输出文件名与原文件名一致。"})
            if default_count >= threshold:
                warning_rows.append({"提醒类型": "大量空白名称使用默认原名", "数量": default_count, "说明": "较多 new_stem 为空，已默认使用原文件名主体。"})
            if no_extension_count >= threshold:
                warning_rows.append({"提醒类型": "大量取消保留扩展名", "数量": no_extension_count, "说明": "较多文件将不保留原扩展名，请确认符合预期。"})
            if contained_ext_count >= threshold:
                warning_rows.append({"提醒类型": "大量填写值包含原扩展名", "数量": contained_ext_count, "说明": "系统已自动去除重复原扩展名，避免生成重复扩展名。"})

        issue_df = pd.DataFrame(issue_rows)
        collision_df = pd.DataFrame(collision_rows)
        warning_df = pd.DataFrame(warning_rows)
        ok = issue_df.empty and collision_df.empty and not plan_df.empty
        return {
            "ok": ok,
            "plan_df": plan_df,
            "issue_df": issue_df,
            "collision_df": collision_df,
            "warning_df": warning_df,
        }

    def _plan_display_df(self, plan_df, max_rows=None):
        if plan_df is None or plan_df.empty:
            return pd.DataFrame()
        display_cols = [
            "原文件名",
            "原扩展名",
            "用户填写的新文件名主体",
            "清洗后的新文件名主体",
            "是否保留原扩展名",
            "最终输出文件名",
            "是否实际改名",
            "处理说明",
            "状态",
        ]
        df = plan_df[display_cols].copy()
        return df.head(max_rows) if max_rows else df

    def _show_validation_success(self, result):
        plan_df = result["plan_df"]
        total = len(plan_df)
        renamed_count = int(plan_df["是否实际改名"].sum())
        unchanged_count = total - renamed_count
        no_extension_count = int((~plan_df["是否保留原扩展名"]).sum())

        summary_html = f"""
        总文件数：<b>{total}</b><br>
        实际改名数量：<b>{renamed_count}</b><br>
        保持原名数量：<b>{unchanged_count}</b><br>
        取消保留扩展名数量：<b>{no_extension_count}</b>
        """

        with self.validation_output:
            clear_output(wait=True)
            display(HTML(panel_html("校验通过，可以执行批量改名", summary_html, "success")))
            display(HTML(dataframe_html(self._plan_display_df(plan_df), title="改名计划预览", max_rows=50)))
            if not result["warning_df"].empty:
                display(HTML(dataframe_html(result["warning_df"], title="提醒信息（不阻断执行）", max_rows=50)))
            local_warnings = self._detect_local_misoperation_warnings(plan_df, self.rename_state)
            if local_warnings:
                display(HTML(dataframe_html(pd.DataFrame(local_warnings), title="本地误操作预警（不阻断执行）", max_rows=80)))

    def _show_validation_failure(self, result, title="校验失败，请修复后重新校验"):
        issue_count = len(result.get("issue_df", pd.DataFrame()))
        collision_count = len(result.get("collision_df", pd.DataFrame()))
        summary_html = f"""
        一般错误数量：<b>{issue_count}</b><br>
        冲突错误数量：<b>{collision_count}</b><br>
        请根据下方明细修改在线表格或批量粘贴内容，然后重新校验。
        """

        with self.validation_output:
            clear_output(wait=True)
            display(HTML(panel_html(title, summary_html, "error")))
            if not result.get("issue_df", pd.DataFrame()).empty:
                display(HTML(dataframe_html(result["issue_df"], title="一般错误明细", max_rows=80)))
            if not result.get("collision_df", pd.DataFrame()).empty:
                display(HTML(dataframe_html(result["collision_df"], title="最终文件名冲突明细", max_rows=80)))
            local_warnings = self._detect_local_misoperation_warnings(result.get("plan_df", pd.DataFrame()), self.rename_state)
            if local_warnings:
                display(HTML(dataframe_html(pd.DataFrame(local_warnings), title="本地误操作预警", max_rows=80)))

    def _report_df(self, plan_df):
        report_cols = [
            "原文件名",
            "最终输出文件名",
            "原扩展名",
            "用户填写的新文件名主体",
            "清洗后的新文件名主体",
            "是否保留原扩展名",
            "是否实际改名",
            "AI建议名",
            "AI建议原因",
            "AI置信度",
            "AI审核结果",
            "AI审核风险",
            "AI审核说明",
            "处理说明",
            "状态",
        ]
        df = plan_df[report_cols].copy()
        df = df.rename(columns={
            "最终输出文件名": "输出文件名",
            "用户填写的新文件名主体": "新文件名主体",
        })
        return df

    def _build_rename_report(self, plan_df):
        report_path = self.main_output_dir / "rename_report.xlsx"
        report_df = self._report_df(plan_df)
        report_df.to_excel(report_path, index=False)
        beautify_excel(report_path)
        return report_path

    def _build_result_zip(self):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        zip_name = f"renamed_files_{timestamp}.zip"
        zip_path = self.export_dir / zip_name
        safe_remove(zip_path)

        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_STORED, allowZip64=True) as zipf:
            for file_path in sorted(self.main_output_dir.iterdir(), key=lambda p: p.name.lower()):
                if file_path.is_file():
                    zipf.write(file_path, arcname=file_path.name)

        return zip_path

    def _execute_rename(self, plan_df):
        safe_remove(self.main_output_dir)
        self.main_output_dir.mkdir(parents=True, exist_ok=True)
        self.export_dir.mkdir(parents=True, exist_ok=True)

        copied_count = 0
        for _, row in plan_df.iterrows():
            src_path = Path(row["__src_path"])
            dst_path = Path(row["__dst_path"])
            if not src_path.exists():
                raise FileNotFoundError(f"源文件不存在：{src_path}")
            if not self._is_inside_dir(dst_path, self.main_output_dir):
                raise RuntimeError(f"目标路径安全校验失败：{dst_path}")
            if dst_path.parent.resolve(strict=False) != self.main_output_dir.resolve(strict=False):
                raise RuntimeError(f"不允许在 output 中生成子目录：{dst_path}")
            shutil.copy2(src_path, dst_path)
            copied_count += 1

        plan_df.loc[:, "状态"] = "已完成"
        report_path = self._build_rename_report(plan_df)
        zip_path = self._build_result_zip()

        self.latest_download_path = zip_path
        self.latest_download_label = zip_path.name
        history_result = {"ok": False, "message": "", "count": 0}
        try:
            append_result = self._append_history_from_plan(plan_df)
            history_result = {
                "ok": True,
                "message": f"历史记录已追加 {append_result['count']} 条，任务批次ID：{append_result['task_id']}",
                "count": append_result["count"],
            }
        except Exception as e:
            history_result = {
                "ok": False,
                "message": f"历史记录写入失败：{e}",
                "count": 0,
            }
            self._log(history_result["message"])
        return {
            "copied_count": copied_count,
            "report_path": report_path,
            "zip_path": zip_path,
            "zip_name": zip_path.name,
            "zip_size": zip_path.stat().st_size,
            "history_result": history_result,
        }

    def _show_execute_success(self, plan_df, result):
        renamed_count = int(plan_df["是否实际改名"].sum())
        zip_size_mb = result["zip_size"] / (1024 * 1024)
        summary_html = f"""
        输出文件数：<b>{result['copied_count']}</b><br>
        实际改名数量：<b>{renamed_count}</b><br>
        ZIP 文件名：<b>{escape(result['zip_name'])}</b><br>
        ZIP 大小：<b>{zip_size_mb:.2f} MB</b><br>
        rename_report.xlsx 已生成并打包进 ZIP。
        """
        history_result = result.get("history_result", {})
        if history_result:
            tone = "success" if history_result.get("ok") else "warning"
            history_html = panel_html("历史记录", escape(history_result.get("message", "")), tone)
        else:
            history_html = ""

        with self.execute_output:
            clear_output(wait=True)
            display(HTML(panel_html("执行成功，结果包已生成", summary_html, "success")))
            if history_html:
                display(HTML(history_html))
            display(HTML(dataframe_html(self._plan_display_df(plan_df), title="改名结果预览", max_rows=50)))

    def _download_file(self, path: Path):
        if path is None or not Path(path).exists():
            raise FileNotFoundError("下载文件不存在。")
        files.download(str(path))

    # --------------------------------------------------------
    # Rule dictionary rename module
    # --------------------------------------------------------
    def _save_uploaded_rule_files(self, uploaded):
        safe_remove(self.rule_upload_dir)
        safe_remove(self.rule_output_dir)
        self.rule_upload_dir.mkdir(parents=True, exist_ok=True)
        self.rule_output_dir.mkdir(parents=True, exist_ok=True)
        saved = []
        for original_name, content in uploaded.items():
            safe_name = Path(original_name).name or "untitled"
            target_path = self.rule_upload_dir / safe_name
            if target_path.exists():
                safe_remove(target_path)
            with open(target_path, "wb") as f:
                f.write(content)
            temp_colab_path = Path("/content") / safe_name
            if temp_colab_path.exists() and temp_colab_path != target_path:
                safe_remove(temp_colab_path)
            saved.append(safe_name)
        return saved

    def _build_rule_rename_plan(self):
        rows = []
        for file_path in sorted(self.rule_upload_dir.iterdir(), key=lambda p: p.name.lower()):
            if not file_path.is_file():
                continue
            stem, extension, _ = split_filename(file_path.name)
            prefix_number = extract_prefix_number(stem)
            hit = prefix_number in DEFAULT_RULE_RENAME_MAP
            if hit:
                target_stem = DEFAULT_RULE_RENAME_MAP[prefix_number]
                process_type = "命中规则"
            else:
                target_stem = re.sub(r"_vlm1_spd1x$", "", stem, flags=re.IGNORECASE)
                target_stem = sanitize_filename_component(target_stem)
                process_type = "未命中规则，保留原主体并清理默认后缀"
            final_name = sanitize_filename_component(target_stem) + extension
            dst_path = self.rule_output_dir / final_name
            rows.append({
                "原文件名": file_path.name,
                "原文件名主体": stem,
                "原扩展名": extension,
                "提取编号": prefix_number,
                "是否命中规则": hit,
                "规则目标主体": target_stem,
                "最终输出文件名": final_name,
                "处理方式": process_type,
                "源文件路径": str(file_path),
                "目标文件路径": str(dst_path),
                "状态": "待执行",
                "__final_key": final_name.lower(),
                "__src_path": str(file_path),
                "__dst_path": str(dst_path),
            })
        return pd.DataFrame(rows)

    def _validate_rule_plan(self, plan_df):
        issues = []
        collisions = []
        if plan_df.empty:
            issues.append({"问题类型": "未上传文件", "原文件名": "", "最终输出文件名": "", "说明": "请上传需要规则处理的文件。"})
        for _, row in plan_df.iterrows():
            final_name = str(row["最终输出文件名"])
            dst_path = Path(row["__dst_path"])
            if final_name.lower() == "rule_rename_report.xlsx":
                issues.append({"问题类型": "输出文件名占用报告文件名", "原文件名": row["原文件名"], "最终输出文件名": final_name, "说明": "rule_rename_report.xlsx 会作为系统报告文件生成。"})
            if re.search(r"[\\/]", final_name) or Path(final_name).name != final_name:
                issues.append({"问题类型": "文件名包含路径分隔符", "原文件名": row["原文件名"], "最终输出文件名": final_name, "说明": "规则模块输出文件名不允许包含目录路径。"})
            if not Path(row["__src_path"]).exists():
                issues.append({"问题类型": "源文件不存在", "原文件名": row["原文件名"], "最终输出文件名": final_name, "说明": row["__src_path"]})
            if not self._is_inside_dir(dst_path, self.rule_output_dir):
                issues.append({"问题类型": "目标路径越界", "原文件名": row["原文件名"], "最终输出文件名": final_name, "说明": str(dst_path)})
        if not plan_df.empty:
            dup_df = plan_df[plan_df["__final_key"].duplicated(keep=False)].copy()
            for _, row in dup_df.sort_values("__final_key").iterrows():
                collisions.append({
                    "原文件名": row["原文件名"],
                    "最终输出文件名": row["最终输出文件名"],
                    "提取编号": row["提取编号"],
                    "处理方式": row["处理方式"],
                })
        return {
            "ok": not issues and not collisions and not plan_df.empty,
            "issue_df": pd.DataFrame(issues),
            "collision_df": pd.DataFrame(collisions),
        }

    def _rule_report_df(self, plan_df):
        df = plan_df[[
            "原文件名",
            "最终输出文件名",
            "原扩展名",
            "提取编号",
            "是否命中规则",
            "规则目标主体",
            "处理方式",
            "状态",
        ]].copy()
        return df.rename(columns={"最终输出文件名": "输出文件名"})

    def _build_rule_report(self, plan_df):
        report_path = self.rule_output_dir / "rule_rename_report.xlsx"
        self._rule_report_df(plan_df).to_excel(report_path, index=False)
        beautify_excel(report_path)
        return report_path

    def _rule_zip_path(self):
        raw_name = str(self.rule_zip_name_text.value or "").strip()
        if raw_name.lower().endswith(".zip"):
            raw_name = raw_name[:-4]
        if raw_name:
            zip_name = sanitize_filename_component(raw_name) + ".zip"
        else:
            zip_name = f"rule_renamed_files_{time.strftime('%Y%m%d_%H%M%S')}.zip"
        return self.export_dir / zip_name

    def _build_rule_zip(self):
        self.export_dir.mkdir(parents=True, exist_ok=True)
        zip_path = self._rule_zip_path()
        safe_remove(zip_path)
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_STORED, allowZip64=True) as zipf:
            for file_path in sorted(self.rule_output_dir.iterdir(), key=lambda p: p.name.lower()):
                if file_path.is_file():
                    zipf.write(file_path, arcname=file_path.name)
        return zip_path

    def _append_rule_history_from_plan(self, plan_df):
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        task_id = f"rule_{time.strftime('%Y%m%d_%H%M%S')}"
        history_df = self._load_history_df(show_error=False)
        new_rows = []
        for _, row in plan_df.iterrows():
            original_stem = row["原文件名主体"]
            final_name = row["最终输出文件名"]
            final_stem, _, _ = split_filename(final_name)
            original_num = extract_prefix_number(original_stem)
            final_num = extract_prefix_number(final_stem)
            new_rows.append({
                "时间戳": timestamp,
                "任务批次ID": task_id,
                "原文件名": row["原文件名"],
                "原文件名主体": original_stem,
                "原扩展名": row["原扩展名"],
                "用户填写的新文件名主体": row["规则目标主体"],
                "清洗后的新文件名主体": final_stem,
                "最终输出文件名": final_name,
                "是否保留原扩展名": True,
                "是否实际改名": row["原文件名"] != final_name,
                "处理说明": row["处理方式"],
                "AI建议名": "",
                "AI建议原因": "",
                "AI置信度": "",
                "AI审核结果": "",
                "AI审核风险": "",
                "AI审核说明": "",
                "状态": "成功",
                "original_stem_norm": normalize_for_history(original_stem),
                "final_stem_norm": normalize_for_history(final_stem),
                "original_prefix_number": original_num,
                "final_prefix_number": final_num,
                "original_tokens": tokenize_for_history(original_stem),
                "final_tokens": tokenize_for_history(final_stem),
                "pattern_key": f"prefix_number_{original_num}" if original_num else "",
            })
        merged_df = pd.concat([history_df, pd.DataFrame(new_rows)], ignore_index=True)
        self._save_history_df(merged_df)
        return {"task_id": task_id, "count": len(new_rows)}

    def _show_rule_failure(self, validation):
        with self.rule_output:
            clear_output(wait=True)
            display(HTML(panel_html("规则处理已拦截", "请根据下方明细修复冲突或问题后重新处理。", "error")))
            if not validation["issue_df"].empty:
                display(HTML(dataframe_html(validation["issue_df"], title="规则模块错误明细", max_rows=80)))
            if not validation["collision_df"].empty:
                display(HTML(dataframe_html(validation["collision_df"], title="规则模块最终文件名冲突明细", max_rows=80)))

    def _execute_rule_rename(self, plan_df):
        safe_remove(self.rule_output_dir)
        self.rule_output_dir.mkdir(parents=True, exist_ok=True)
        copied = 0
        for _, row in plan_df.iterrows():
            src_path = Path(row["__src_path"])
            dst_path = Path(row["__dst_path"])
            if not src_path.exists():
                raise FileNotFoundError(f"源文件不存在：{src_path}")
            if dst_path.parent.resolve(strict=False) != self.rule_output_dir.resolve(strict=False):
                raise RuntimeError(f"不允许在规则 output 中生成子目录：{dst_path}")
            shutil.copy2(src_path, dst_path)
            copied += 1
        plan_df.loc[:, "状态"] = "成功"
        self._build_rule_report(plan_df)
        zip_path = self._build_rule_zip()
        self.latest_rule_download_path = zip_path
        self.latest_rule_download_label = zip_path.name
        history_message = ""
        try:
            result = self._append_rule_history_from_plan(plan_df)
            history_message = f"规则模块历史记录已追加 {result['count']} 条，任务批次ID：{result['task_id']}"
        except Exception as e:
            history_message = f"规则模块历史记录写入失败：{e}"
            self._log(history_message)
        return {
            "copied": copied,
            "zip_path": zip_path,
            "zip_name": zip_path.name,
            "zip_size": zip_path.stat().st_size,
            "history_message": history_message,
        }

    def _on_process_rule_files_clicked(self, _):
        if self.busy:
            return
        self.busy = True
        self._refresh_buttons()
        try:
            self._set_status("等待上传规则处理文件", "info")
            uploaded = files.upload()
            if not uploaded:
                with self.rule_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("未上传规则处理文件", "请选择需要按规则处理的文件。", "warning")))
                return
            saved = self._save_uploaded_rule_files(uploaded)
            del uploaded
            gc.collect()
            plan_df = self._build_rule_rename_plan()
            validation = self._validate_rule_plan(plan_df)
            if not validation["ok"]:
                self._show_rule_failure(validation)
                return
            result = self._execute_rule_rename(plan_df)
            zip_size_mb = result["zip_size"] / (1024 * 1024)
            summary = f"""
            上传文件数：<b>{len(saved)}</b><br>
            输出文件数：<b>{result['copied']}</b><br>
            ZIP 文件名：<b>{escape(result['zip_name'])}</b><br>
            ZIP 大小：<b>{zip_size_mb:.2f} MB</b><br>
            rule_rename_report.xlsx 已生成并打包进 ZIP。<br>
            {escape(result['history_message'])}
            """
            with self.rule_output:
                clear_output(wait=True)
                display(HTML(panel_html("规则字典自动改名完成", summary, "success")))
                display(HTML(dataframe_html(self._rule_report_df(plan_df), title="规则处理结果预览", max_rows=80)))
            self._show_history_review(mode="recent")
            self._set_status("规则字典自动改名完成，ZIP 下载已触发", "success")
            try:
                self._download_file(result["zip_path"])
            except Exception as e:
                with self.rule_output:
                    display(HTML(panel_html("规则模块下载触发失败", escape(e), "error")))
        except Exception as e:
            with self.rule_output:
                clear_output(wait=True)
                display(HTML(panel_html("规则字典自动改名失败", escape(e), "error")))
            self._log(f"规则字典自动改名失败：{e}")
        finally:
            self.busy = False
            self._refresh_buttons()

    def _on_redownload_rule_latest_clicked(self, _):
        try:
            if self.latest_rule_download_path is None or not Path(self.latest_rule_download_path).exists():
                with self.rule_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("暂无规则模块结果包", "当前没有可重新下载的规则模块结果包。", "warning")))
                return
            self._download_file(self.latest_rule_download_path)
            with self.rule_output:
                display(HTML(panel_html("重新下载规则模块结果包", f"已触发下载：<b>{escape(self.latest_rule_download_label)}</b>", "success")))
        except Exception as e:
            with self.rule_output:
                clear_output(wait=True)
                display(HTML(panel_html("重新下载规则模块结果包失败", escape(e), "error")))

    # --------------------------------------------------------
    # Button handlers
    # --------------------------------------------------------
    def _on_validate_clicked(self, _):
        if self.busy:
            return
        if not self.upload_completed:
            self._show_main_error("无法校验", "未上传文件，请先上传待改名文件。")
            return

        self.busy = True
        self.validation_passed = False
        self.latest_valid_plan_df = None
        self._refresh_buttons()

        try:
            result = self._validate_current_state()
            self.latest_validation_result = result
            if result["ok"]:
                self.validation_passed = True
                self.is_dirty = False
                self.latest_valid_plan_df = result["plan_df"].copy()
                self._set_stage("validation_passed")
                self._set_status("校验通过，可以执行批量改名", "success")
                self._set_progress(85, "校验通过，可以执行批量改名", "success")
                self._show_validation_success(result)
                self._log(f"校验通过：共 {len(result['plan_df'])} 个文件。")
            else:
                self.validation_passed = False
                self.is_dirty = True
                self._set_stage("validation_failed")
                self._set_status("校验失败，请修复后重新校验", "error")
                self._set_progress(70, "校验失败，请修复后重新校验", "error")
                self._show_validation_failure(result)
                self._log("校验失败，已展示错误明细。")
        except Exception as e:
            self.validation_passed = False
            self.latest_valid_plan_df = None
            self._show_main_error("校验异常", str(e))
            self._set_progress(70, "校验异常", "error")
            self._log(f"校验异常：{e}")
        finally:
            self.busy = False
            self._refresh_buttons()

    def _on_execute_clicked(self, _):
        if self.busy:
            return
        if not self.validation_passed or self.latest_valid_plan_df is None:
            self._show_main_error("无法执行", "当前命名尚未通过校验，请先点击“校验当前命名”。")
            return

        self.busy = True
        self._set_stage("executing")
        self._refresh_buttons()

        try:
            self._set_status("正在执行批量改名并生成结果包", "info")
            self._set_progress(90, "正在复制文件并生成报告", "info")

            validation_result = self._validate_current_state()
            if not validation_result["ok"]:
                self.validation_passed = False
                self.latest_valid_plan_df = None
                self._set_status("执行前校验失败，请修复后重新校验", "error")
                self._set_progress(70, "执行前校验失败", "error")
                self._show_validation_failure(validation_result, title="执行前校验失败")
                return

            plan_df = validation_result["plan_df"].copy()
            result = self._execute_rename(plan_df)
            self.latest_valid_plan_df = plan_df
            self.validation_passed = True
            self.is_dirty = False
            self._set_stage("executed")

            self._show_execute_success(plan_df, result)
            self._show_history_review(mode="recent")
            self._set_status("执行成功，ZIP 下载已触发", "success")
            self._set_progress(100, "执行成功，ZIP 下载已触发", "success")
            self._log(f"执行成功：{result['zip_name']}，输出 {result['copied_count']} 个文件。")

            try:
                self._download_file(result["zip_path"])
            except Exception as download_error:
                self._show_main_error("下载触发失败", str(download_error))
                self._log(f"下载触发失败：{download_error}")

        except Exception as e:
            self._set_stage("error")
            self._show_main_error("执行失败", str(e))
            self._set_progress(90, "执行失败", "error")
            self._log(f"执行失败：{e}")
        finally:
            self.busy = False
            self._refresh_buttons()

    def _on_redownload_latest_clicked(self, _):
        if self.busy:
            return
        try:
            if self.latest_download_path is None or not Path(self.latest_download_path).exists():
                with self.execute_output:
                    clear_output(wait=True)
                    display(HTML(panel_html("重新下载失败", "当前没有可重新下载的结果包。", "warning")))
                self._set_status("当前没有可重新下载的结果包", "warning")
                return

            self._download_file(Path(self.latest_download_path))
            self._set_status("最近结果包下载已触发", "success")
            with self.execute_output:
                display(HTML(panel_html(
                    "重新下载最近结果包",
                    f"已触发下载：<b>{escape(self.latest_download_label)}</b>",
                    "success",
                )))
            self._log(f"重新下载最近结果包：{self.latest_download_label}")
        except Exception as e:
            self._show_main_error("重新下载失败", str(e))
            self._log(f"重新下载失败：{e}")

    def _on_use_original_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        for item in self.file_inventory:
            original_name = item["original_name"]
            self.rename_state[original_name]["new_stem"] = sanitize_filename_component(item["stem"])
            self.rename_state[original_name]["keep_extension"] = True
            self.rename_state[original_name]["status"] = "待编辑"
        self._refresh_online_table_from_state()
        self._mark_dirty()
        self._log("已一键恢复所有新文件名主体为清洗后的原文件名主体。")

    def _on_clear_selection_clicked(self, _):
        if not self.upload_completed:
            return
        self._read_online_table_to_state()
        for state in self.rename_state.values():
            state["selected"] = False
            state["use_ai_suggestion"] = False
        self._refresh_online_table_from_state()
        self._mark_dirty()
        self._log("已清空所有选择和采用AI建议状态。")

    def _on_reset_clicked(self, _):
        if self.busy:
            return
        self.busy = True
        self._refresh_buttons()
        try:
            self._start_new_task()
            self._set_status("已重置，可开始新任务", "success")
            self._set_progress(0, "已重置，可开始新任务", "success")
            self._log("已结束当前任务并开启新任务。")
        except Exception as e:
            self._show_main_error("重置失败", str(e))
        finally:
            self.busy = False
            self._refresh_buttons()

    def _on_rule_placeholder_clicked(self, _):
        self._set_status("规则字典自动改名模块已可使用", "info")
        with self.rule_output:
            clear_output(wait=True)
            display(HTML(panel_html(
                "规则字典自动改名模块",
                "规则字典自动改名模块已完整接入，可在本区域上传任意文件并按开头编号自动处理；兼容原编号规则。",
                "info",
            )))
        self._log("用户打开规则字典自动改名模块。")


# ============================================================
# 启动应用
# ============================================================
def launch_colab_app():
    """启动 Colab 版 ipywidgets 应用。"""
    app = UniversalColabRenameApp()
    app.show()
    return app


if __name__ == "__main__":
    launch_colab_app()
